from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import wave
import zipfile
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext as _
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook

from .forms import (
    BehaviorCategoryForm,
    BehaviorForm,
    DeleteConfirmForm,
    EthogramImportForm,
    IndependentVariableDefinitionForm,
    KeyboardProfileForm,
    ModifierForm,
    ObservationSegmentForm,
    ObservationSessionForm,
    ObservationTemplateForm,
    ProjectBORISImportForm,
    ProjectCloneForm,
    ProjectForm,
    ProjectImportCreateForm,
    ProjectMembershipForm,
    ProjectSettingsForm,
    SessionImportForm,
    SignUpForm,
    SubjectForm,
    SubjectGroupForm,
    VideoAssetForm,
)
from .models import (
    Behavior,
    BehaviorCategory,
    IndependentVariableDefinition,
    KeyboardProfile,
    Modifier,
    ObservationAuditLog,
    ObservationEvent,
    ObservationSegment,
    ObservationSession,
    ObservationTemplate,
    ObservationVariableValue,
    Project,
    ProjectMembership,
    SessionAnnotation,
    SessionVideoLink,
    Subject,
    SubjectGroup,
    VideoAsset,
)


def signup(request):  # pragma: no cover
    if request.user.is_authenticated:
        return redirect('tracker:home')

    form = SignUpForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, _('Account created successfully.'))
        return redirect('tracker:home')
    return render(request, 'registration/signup.html', {'form': form})


def accessible_projects_qs(user):
    """Return every project visible to the current authenticated user."""
    return (
        Project.objects.filter(Q(owner=user) | Q(collaborators=user) | Q(memberships__user=user))
        .distinct()
        .select_related('owner')
        .prefetch_related('collaborators', 'memberships__user', 'keyboard_profiles')
    )


def get_accessible_project(user, pk: int) -> Project:
    return get_object_or_404(accessible_projects_qs(user), pk=pk)


def project_role(user, project: Project) -> str | None:
    return project.role_for_user(user)


def _require_project_owner(user, project: Project, message: str | None = None) -> None:
    if project_role(user, project) != ProjectMembership.ROLE_OWNER:
        raise PermissionDenied(message or _('Only the project owner can update project settings.'))


def _require_project_editor(user, project: Project, message: str | None = None) -> None:
    if not project.can_edit(user):
        raise PermissionDenied(message or _('You need editor permissions for this project.'))


def _require_project_reviewer(user, project: Project, message: str | None = None) -> None:
    if not project.can_review(user):
        raise PermissionDenied(message or _('You need reviewer permissions for this project.'))


def get_owned_project(user, pk: int) -> Project:
    project = get_accessible_project(user, pk)
    _require_project_owner(user, project)
    return project


def build_release_metadata() -> dict:
    """Return a small machine-readable release description for health and ops tooling."""
    return {
        'application': 'PyBehaviorLog',
        'version': '0.9.2',
        'django_target': '6.0.3',
        'python_minimum': '3.13',
        'asgi': True,
        'server': 'granian',
        'database': 'postgresql 18',
        'cache': 'redis 8',
        'languages': ['en', 'ar', 'zh-hans', 'es', 'fr', 'ru'],
    }


def clone_project(  # pragma: no cover
    source: Project,
    *,
    owner,
    name: str,
    description: str = '',
    include_sessions: bool = True,
    include_videos: bool = True,
) -> Project:
    """Clone a project and its coding resources into a new owner-owned project."""
    with transaction.atomic():
        cloned = Project.objects.create(owner=owner, name=name, description=description)
        ProjectMembership.objects.update_or_create(
            project=cloned,
            user=owner,
            defaults={'role': ProjectMembership.ROLE_OWNER},
        )

        category_map = {}
        for category in source.categories.order_by('sort_order', 'name'):
            category_map[category.pk] = BehaviorCategory.objects.create(
                project=cloned,
                name=category.name,
                color=category.color,
                sort_order=category.sort_order,
            )

        modifier_map = {}
        for modifier in source.modifiers.order_by('sort_order', 'name'):
            modifier_map[modifier.pk] = Modifier.objects.create(
                project=cloned,
                name=modifier.name,
                description=modifier.description,
                key_binding=modifier.key_binding,
                sort_order=modifier.sort_order,
            )

        group_map = {}
        for group in source.subject_groups.order_by('sort_order', 'name'):
            group_map[group.pk] = SubjectGroup.objects.create(
                project=cloned,
                name=group.name,
                description=group.description,
                color=group.color,
                sort_order=group.sort_order,
            )

        subject_map = {}
        for subject in source.subjects.order_by('sort_order', 'name').prefetch_related('groups'):
            new_subject = Subject.objects.create(
                project=cloned,
                name=subject.name,
                description=subject.description,
                key_binding=subject.key_binding,
                color=subject.color,
                sort_order=subject.sort_order,
            )
            new_subject.groups.set(
                [group_map[group.pk] for group in subject.groups.all() if group.pk in group_map]
            )
            subject_map[subject.pk] = new_subject

        variable_map = {}
        for variable in source.variable_definitions.order_by('sort_order', 'label'):
            variable_map[variable.pk] = IndependentVariableDefinition.objects.create(
                project=cloned,
                label=variable.label,
                description=variable.description,
                value_type=variable.value_type,
                set_values=variable.set_values,
                default_value=variable.default_value,
                sort_order=variable.sort_order,
            )

        behavior_map = {}
        for behavior in source.behaviors.order_by('sort_order', 'name').select_related('category'):
            behavior_map[behavior.pk] = Behavior.objects.create(
                project=cloned,
                category=category_map.get(behavior.category_id),
                name=behavior.name,
                description=behavior.description,
                key_binding=behavior.key_binding,
                color=behavior.color,
                mode=behavior.mode,
                sort_order=behavior.sort_order,
            )

        for template in source.observation_templates.order_by('name').prefetch_related(
            'behaviors', 'modifiers', 'subjects', 'variable_definitions'
        ):
            new_template = ObservationTemplate.objects.create(
                project=cloned,
                name=template.name,
                description=template.description,
                default_session_kind=template.default_session_kind,
            )
            new_template.behaviors.set(
                [
                    behavior_map[item.pk]
                    for item in template.behaviors.all()
                    if item.pk in behavior_map
                ]
            )
            new_template.modifiers.set(
                [
                    modifier_map[item.pk]
                    for item in template.modifiers.all()
                    if item.pk in modifier_map
                ]
            )
            new_template.subjects.set(
                [subject_map[item.pk] for item in template.subjects.all() if item.pk in subject_map]
            )
            new_template.variable_definitions.set(
                [
                    variable_map[item.pk]
                    for item in template.variable_definitions.all()
                    if item.pk in variable_map
                ]
            )

        profile_map = {}
        for profile in source.keyboard_profiles.order_by('name'):
            profile_map[profile.pk] = KeyboardProfile.objects.create(
                project=cloned,
                name=profile.name,
                description=profile.description,
                is_default=profile.is_default,
                behavior_bindings=profile.behavior_bindings,
                modifier_bindings=profile.modifier_bindings,
                subject_bindings=profile.subject_bindings,
            )

        video_map = {}
        if include_videos:
            for video in source.videos.order_by('title'):
                new_video = VideoAsset.objects.create(
                    project=cloned,
                    title=video.title,
                    file=video.file.name,
                    notes=video.notes,
                )
                video_map[video.pk] = new_video

        if include_sessions:
            template_map = {item.name: item for item in cloned.observation_templates.all()}
            session_video_links = []
            annotation_rows = []
            segment_rows = []
            session_map = {}
            m2m_event_modifiers = []
            m2m_event_subjects = []
            for session in (
                source.sessions.order_by('created_at')
                .select_related('video', 'template', 'keyboard_profile', 'observer')
                .prefetch_related(
                    'events__modifiers',
                    'events__subjects',
                    'annotations',
                    'segments',
                    'video_links__video',
                    'variable_values__definition',
                )
            ):
                new_session = ObservationSession.objects.create(
                    project=cloned,
                    video=video_map.get(session.video_id),
                    template=template_map.get(session.template.name)
                    if session.template_id
                    else None,
                    keyboard_profile=profile_map.get(session.keyboard_profile_id),
                    session_kind=session.session_kind,
                    workflow_status=session.workflow_status,
                    title=session.title,
                    description=session.description,
                    observer=owner,
                    notes=session.notes,
                    review_notes=session.review_notes,
                    reviewed_by=None,
                    reviewed_at=session.reviewed_at,
                    locked_at=session.locked_at,
                    playback_rate=session.playback_rate,
                    frame_step_seconds=session.frame_step_seconds,
                    recorded_at=session.recorded_at,
                )
                session_map[session.pk] = new_session
                for value in session.variable_values.all():
                    if value.definition_id in variable_map:
                        ObservationVariableValue.objects.create(
                            session=new_session,
                            definition=variable_map[value.definition_id],
                            value=value.value,
                        )
                for link in session.video_links.all():
                    if link.video_id in video_map:
                        session_video_links.append(
                            SessionVideoLink(
                                session=new_session,
                                video=video_map[link.video_id],
                                sort_order=link.sort_order,
                            )
                        )
                event_map = {}
                for event in session.events.all():
                    new_event = ObservationEvent.objects.create(
                        session=new_session,
                        subject=subject_map.get(event.subject_id),
                        behavior=behavior_map[event.behavior_id],
                        event_kind=event.event_kind,
                        timestamp_seconds=event.timestamp_seconds,
                        frame_index=event.frame_index,
                        comment=event.comment,
                    )
                    event_map[event.pk] = new_event
                    m2m_event_modifiers.append(
                        (
                            new_event,
                            [
                                modifier_map[item.pk]
                                for item in event.modifiers.all()
                                if item.pk in modifier_map
                            ],
                        )
                    )
                    m2m_event_subjects.append(
                        (
                            new_event,
                            [
                                subject_map[item.pk]
                                for item in event.subjects.all()
                                if item.pk in subject_map
                            ],
                        )
                    )
                for annotation in session.annotations.all():
                    annotation_rows.append(
                        SessionAnnotation(
                            session=new_session,
                            timestamp_seconds=annotation.timestamp_seconds,
                            title=annotation.title,
                            note=annotation.note,
                            color=annotation.color,
                            created_by=owner,
                        )
                    )
                for segment in session.segments.all():
                    segment_rows.append(
                        ObservationSegment(
                            session=new_session,
                            title=segment.title,
                            start_seconds=segment.start_seconds,
                            end_seconds=segment.end_seconds,
                            status=segment.status,
                            assignee=owner if segment.assignee_id else None,
                            reviewer=owner if segment.reviewer_id else None,
                            notes=segment.notes,
                        )
                    )
            if session_video_links:
                SessionVideoLink.objects.bulk_create(session_video_links)
            if annotation_rows:
                SessionAnnotation.objects.bulk_create(annotation_rows)
            if segment_rows:
                ObservationSegment.objects.bulk_create(segment_rows)
            for event, modifiers in m2m_event_modifiers:
                if modifiers:
                    event.modifiers.set(modifiers)
            for event, subjects in m2m_event_subjects:
                if subjects:
                    event.subjects.set(subjects)

        return cloned


def accessible_sessions_qs(user):
    return (
        ObservationSession.objects.filter(
            Q(project__owner=user)
            | Q(project__collaborators=user)
            | Q(project__memberships__user=user)
        )
        .distinct()
        .select_related('project', 'video', 'observer', 'project__owner', 'keyboard_profile')
    )


def get_accessible_session(user, pk: int) -> ObservationSession:
    session_qs = accessible_sessions_qs(user).prefetch_related(
        'project__behaviors__category',
        'project__modifiers',
        'project__categories',
        'project__subjects__groups',
        'project__subject_groups',
        'project__variable_definitions',
        'project__observation_templates',
        'project__keyboard_profiles',
        'video_links__video',
        'variable_values__definition',
        'audit_logs__actor',
        Prefetch(
            'events',
            queryset=ObservationEvent.objects.select_related(
                'behavior', 'behavior__category', 'subject'
            )
            .prefetch_related('modifiers', 'subjects')
            .order_by('timestamp_seconds', 'pk'),
        ),
        'annotations',
    )
    return get_object_or_404(session_qs, pk=pk)


def build_review_queue(user) -> dict:
    projects = accessible_projects_qs(user)
    segments = list(
        ObservationSegment.objects.filter(session__project__in=projects)
        .select_related('session', 'session__project', 'assignee', 'reviewer')
        .order_by('status', 'session__project__name', 'session__title', 'start_seconds')
    )
    outstanding = [
        segment for segment in segments if segment.status != ObservationSegment.STATUS_DONE
    ]
    assigned = [segment for segment in outstanding if segment.assignee_id == user.id]
    review = [segment for segment in outstanding if segment.reviewer_id == user.id]
    return {
        'all': segments,
        'outstanding': outstanding,
        'assigned': assigned,
        'review': review,
        'counts': {
            'all': len(segments),
            'outstanding': len(outstanding),
            'assigned': len(assigned),
            'review': len(review),
        },
    }


def _get_owned_category(user, pk: int) -> BehaviorCategory:
    category = get_object_or_404(BehaviorCategory.objects.select_related('project'), pk=pk)
    _require_project_editor(
        user, category.project, _('You need editor permissions to manage categories.')
    )
    return category


def _get_owned_modifier(user, pk: int) -> Modifier:
    modifier = get_object_or_404(Modifier.objects.select_related('project'), pk=pk)
    _require_project_editor(
        user, modifier.project, _('You need editor permissions to manage modifiers.')
    )
    return modifier


def _get_owned_behavior(user, pk: int) -> Behavior:
    behavior = get_object_or_404(Behavior.objects.select_related('project', 'category'), pk=pk)
    _require_project_editor(
        user, behavior.project, _('You need editor permissions to manage behaviors.')
    )
    return behavior


def _get_owned_video(user, pk: int) -> VideoAsset:
    video = get_object_or_404(VideoAsset.objects.select_related('project'), pk=pk)
    _require_project_editor(user, video.project, _('You need editor permissions to manage videos.'))
    return video


def _require_editable_session(session: ObservationSession, user=None) -> None:
    if user is not None and not session.project.can_edit(user):
        raise PermissionDenied(_('You need editor permissions to modify this session.'))
    if session.is_locked_for_coding:
        raise PermissionDenied(_('This session is locked and cannot be modified.'))


def _log_audit(
    session: ObservationSession,
    *,
    actor,
    action: str,
    target_type: str,
    summary: str,
    payload: dict | None = None,
    target_id: int | None = None,
) -> ObservationAuditLog:
    return ObservationAuditLog.objects.create(
        session=session,
        actor=actor,
        action=action,
        target_type=target_type,
        target_id=target_id,
        summary=summary,
        payload=payload or {},
    )


def _decimal(value, default: str = '0') -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _session_duration(
    session: ObservationSession, duration_hint: str | float | Decimal | None = None
) -> Decimal:
    duration = _decimal(duration_hint, default='0')
    if duration > 0:
        return duration
    last_event = (
        ObservationEvent.objects.filter(session=session).order_by('-timestamp_seconds').first()
    )
    last_annotation = (
        SessionAnnotation.objects.filter(session=session).order_by('-timestamp_seconds').first()
    )
    candidates = [Decimal('0')]
    if last_event is not None:
        candidates.append(last_event.timestamp_seconds)
    if last_annotation is not None:
        candidates.append(last_annotation.timestamp_seconds)
    return max(candidates)


def _relative_media_path(video: VideoAsset | None) -> str | None:
    """Return a storage-relative media path for interoperability exports."""
    if video is None or not getattr(video, 'file', None):
        return None
    name = str(video.file.name or '').replace('\\', '/')
    return name or None


def _resolve_storage_path(video: VideoAsset | None) -> Path | None:
    """Resolve a local filesystem path when the file is available on local storage."""
    if video is None or not getattr(video, 'file', None):
        return None
    candidate = getattr(video.file, 'path', None)
    if not candidate:
        return None
    try:
        return Path(candidate)
    except (TypeError, ValueError):
        return None


def _media_kind_from_name(name: str | None) -> str:
    suffix = Path(name or '').suffix.lower()
    if suffix in {'.wav', '.mp3', '.flac', '.ogg', '.m4a'}:
        return 'audio'
    if suffix in {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tif', '.tiff'}:
        return 'image'
    if suffix in {'.mp4', '.webm', '.mov', '.avi', '.mkv', '.m4v'}:
        return 'video'
    return 'file'


def _downsample(values: list[float], target_points: int) -> list[float]:
    if len(values) <= target_points:
        return values
    bucket_size = max(len(values) // target_points, 1)
    results: list[float] = []
    for start in range(0, len(values), bucket_size):
        chunk = values[start : start + bucket_size]
        if not chunk:
            continue
        results.append(round(sum(chunk) / len(chunk), 6))
    return results[:target_points]


def _wav_visual_summary(
    file_path: Path, *, points: int = 96, spectrogram_columns: int = 24, spectrogram_rows: int = 8
) -> dict:
    """Build a lightweight waveform and coarse spectrogram for WAV files using stdlib only."""
    try:
        with wave.open(str(file_path), 'rb') as wav_file:
            channels = max(wav_file.getnchannels(), 1)
            sample_width = wav_file.getsampwidth()
            frame_rate = wav_file.getframerate() or 1
            frame_count = wav_file.getnframes()
            raw_frames = wav_file.readframes(frame_count)
    except (wave.Error, OSError):
        return {'available': False, 'reason': 'unreadable'}

    if sample_width not in {1, 2, 4} or not raw_frames:
        return {'available': False, 'reason': 'unsupported'}

    step = sample_width * channels
    samples: list[float] = []
    for index in range(0, len(raw_frames), step):
        chunk = raw_frames[index : index + sample_width]
        if len(chunk) < sample_width:
            break
        signed = sample_width != 1
        value = int.from_bytes(chunk, byteorder='little', signed=signed)
        if sample_width == 1:
            value -= 128
            scale = 128.0
        else:
            scale = float(2 ** (sample_width * 8 - 1))
        samples.append(max(min(value / scale, 1.0), -1.0))

    if not samples:
        return {'available': False, 'reason': 'empty'}

    waveform = _downsample([abs(sample) for sample in samples], points)

    # Coarse spectrogram using a tiny DFT over evenly spaced windows.
    window_size = min(256, max(64, len(samples) // max(spectrogram_columns, 1) or 64))
    spectrogram: list[list[float]] = []
    if window_size < 8:
        window_size = min(len(samples), 8)
    max_start = max(len(samples) - window_size, 0)
    column_starts = [
        int(round(max_start * index / max(spectrogram_columns - 1, 1)))
        for index in range(spectrogram_columns)
    ]
    frequency_bins = [
        1 + index * max(window_size // (2 * spectrogram_rows), 1)
        for index in range(spectrogram_rows)
    ]
    for start in column_starts:
        window = samples[start : start + window_size]
        if len(window) < window_size:
            window = window + [0.0] * (window_size - len(window))
        column: list[float] = []
        for bin_index in frequency_bins:
            real = 0.0
            imag = 0.0
            for sample_index, sample in enumerate(window):
                angle = 2 * math.pi * bin_index * sample_index / max(window_size, 1)
                real += sample * math.cos(angle)
                imag -= sample * math.sin(angle)
            magnitude = (real * real + imag * imag) ** 0.5 / max(window_size, 1)
            column.append(round(magnitude, 6))
        peak = max(column, default=1.0) or 1.0
        spectrogram.append([round(value / peak, 6) for value in column])

    return {
        'available': True,
        'duration_seconds': round(frame_count / frame_rate, 3),
        'frame_rate': frame_rate,
        'channels': channels,
        'waveform': waveform,
        'spectrogram': spectrogram,
    }


def _image_sequence_summary(file_path: Path | None, *, limit: int = 12) -> dict:
    """Describe a likely image sequence around a local image file."""
    if file_path is None or not file_path.exists() or not file_path.is_file():
        return {'available': False, 'reason': 'missing'}
    suffix = file_path.suffix.lower()
    if suffix not in {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tif', '.tiff'}:
        return {'available': False, 'reason': 'not-image'}
    parent = file_path.parent
    stem = file_path.stem
    prefix = re.sub(r'\d+$', '', stem)
    pattern = (
        re.compile(rf'^{re.escape(prefix)}\d*{re.escape(suffix)}$', re.IGNORECASE)
        if prefix
        else None
    )
    siblings = []
    for candidate in sorted(parent.iterdir()):
        if not candidate.is_file():
            continue
        if candidate.suffix.lower() != suffix:
            continue
        if pattern is not None and not pattern.match(candidate.name):
            continue
        siblings.append(candidate.name)
    if file_path.name not in siblings:
        siblings.insert(0, file_path.name)
    index = siblings.index(file_path.name) if file_path.name in siblings else 0
    start = max(index - limit // 2, 0)
    end = min(start + limit, len(siblings))
    preview = siblings[start:end]
    return {
        'available': True,
        'directory': str(parent),
        'relative_directory': parent.name,
        'sequence_count': len(siblings),
        'current_index': index,
        'preview_files': preview,
    }


def build_media_analysis(session: ObservationSession) -> list[dict]:
    """Return media diagnostics for synced sources, including relative paths and audio summaries."""
    rows: list[dict] = []
    for video in session.all_videos_ordered:
        relative_path = _relative_media_path(video)
        storage_path = _resolve_storage_path(video)
        item = {
            'id': video.id,
            'title': video.title,
            'relative_path': relative_path,
            'media_kind': _media_kind_from_name(relative_path),
            'file_exists': bool(storage_path and storage_path.exists()),
            'size_bytes': storage_path.stat().st_size
            if storage_path and storage_path.exists()
            else None,
            'waveform': [],
            'spectrogram': [],
            'audio_summary': {'available': False, 'reason': 'not-audio'},
            'image_sequence': {'available': False, 'reason': 'not-image'},
        }
        if (
            item['media_kind'] == 'audio'
            and storage_path
            and storage_path.exists()
            and storage_path.suffix.lower() == '.wav'
        ):
            audio_summary = _wav_visual_summary(storage_path)
            item['audio_summary'] = audio_summary
            item['waveform'] = audio_summary.get('waveform', [])
            item['spectrogram'] = audio_summary.get('spectrogram', [])
        if item['media_kind'] == 'image':
            item['image_sequence'] = _image_sequence_summary(storage_path)
        rows.append(item)
    return rows


def _server_history_key(session_id: int) -> str:
    return f'pybehaviorlog_history_{session_id}'


def _history_event_payload(payload: dict | None) -> dict | None:
    if not isinstance(payload, dict):
        return None
    if isinstance(payload.get('event'), dict):
        return payload['event']
    if 'behavior' in payload and 'timestamp_seconds' in payload:
        return payload
    return None


def _push_server_history(request, session_id: int, entry: dict) -> None:
    key = _server_history_key(session_id)
    stacks = request.session.get(key, {'undo': [], 'redo': []})
    undo_stack = list(stacks.get('undo', []))
    undo_stack.append(entry)
    stacks['undo'] = undo_stack[-50:]
    stacks['redo'] = []
    request.session[key] = stacks
    request.session.modified = True


def _pop_server_history(request, session_id: int, stack_name: str) -> dict | None:
    key = _server_history_key(session_id)
    stacks = request.session.get(key, {'undo': [], 'redo': []})
    stack = list(stacks.get(stack_name, []))
    if not stack:
        return None
    entry = stack.pop()
    stacks[stack_name] = stack
    request.session[key] = stacks
    request.session.modified = True
    return entry


def _push_redo_history(request, session_id: int, entry: dict) -> None:
    key = _server_history_key(session_id)
    stacks = request.session.get(key, {'undo': [], 'redo': []})
    redo_stack = list(stacks.get('redo', []))
    redo_stack.append(entry)
    stacks['redo'] = redo_stack[-50:]
    request.session[key] = stacks
    request.session.modified = True


def _restore_history_entry(request, session_id: int, entry: dict, *, to_stack: str) -> None:
    key = _server_history_key(session_id)
    stacks = request.session.get(key, {'undo': [], 'redo': []})
    target_stack = list(stacks.get(to_stack, []))
    target_stack.append(entry)
    stacks[to_stack] = target_stack[-50:]
    request.session[key] = stacks
    request.session.modified = True


def _create_event_from_snapshot(session: ObservationSession, snapshot: dict) -> ObservationEvent:
    behavior = get_object_or_404(Behavior, project=session.project, pk=snapshot.get('behavior_id'))
    event = ObservationEvent.objects.create(
        session=session,
        behavior=behavior,
        event_kind=snapshot.get('event_kind') or ObservationEvent.KIND_POINT,
        timestamp_seconds=_decimal(snapshot.get('timestamp_seconds'), default='0'),
        frame_index=snapshot.get('frame_index') or None,
        comment=(snapshot.get('comment') or '').strip(),
        subject_id=snapshot.get('subject_id') or None,
    )
    modifier_ids = [
        int(item['id'])
        for item in snapshot.get('modifiers', [])
        if isinstance(item, dict) and item.get('id')
    ]
    subject_ids = [
        int(item['id'])
        for item in snapshot.get('subjects', [])
        if isinstance(item, dict) and item.get('id')
    ]
    if modifier_ids:
        event.modifiers.set(Modifier.objects.filter(project=session.project, pk__in=modifier_ids))
    if subject_ids:
        subjects = list(Subject.objects.filter(project=session.project, pk__in=subject_ids))
        event.subjects.set(subjects)
        if not event.subject_id and subjects:
            event.subject = subjects[0]
            event.save(update_fields=['subject'])
    return event


def _apply_history_entry(session: ObservationSession, entry: dict, *, direction: str) -> str:
    action = entry.get('action')
    before = _history_event_payload(entry.get('before'))
    after = _history_event_payload(entry.get('after'))
    if action == ObservationAuditLog.ACTION_CREATE:
        if direction == 'undo' and after:
            ObservationEvent.objects.filter(session=session, pk=after.get('id')).delete()
            return 'create'
        if direction == 'redo' and after:
            recreated = _create_event_from_snapshot(session, after)
            after['id'] = recreated.id
            return 'create'
    if action == ObservationAuditLog.ACTION_DELETE:
        if direction == 'undo' and before:
            recreated = _create_event_from_snapshot(session, before)
            before['id'] = recreated.id
            return 'delete'
        if direction == 'redo' and before:
            ObservationEvent.objects.filter(session=session, pk=before.get('id')).delete()
            return 'delete'
    if action == ObservationAuditLog.ACTION_UPDATE and before and after:
        snapshot = before if direction == 'undo' else after
        event = get_object_or_404(
            ObservationEvent, session=session, pk=(after.get('id') or before.get('id'))
        )
        behavior = get_object_or_404(
            Behavior, project=session.project, pk=snapshot.get('behavior_id')
        )
        event.behavior = behavior
        event.event_kind = snapshot.get('event_kind') or event.event_kind
        event.timestamp_seconds = _decimal(snapshot.get('timestamp_seconds'), default='0')
        event.frame_index = snapshot.get('frame_index') or None
        event.comment = (snapshot.get('comment') or '').strip()
        event.subject_id = snapshot.get('subject_id') or None
        event.save(
            update_fields=[
                'behavior',
                'event_kind',
                'timestamp_seconds',
                'frame_index',
                'comment',
                'subject',
            ]
        )
        modifier_ids = [
            int(item['id'])
            for item in snapshot.get('modifiers', [])
            if isinstance(item, dict) and item.get('id')
        ]
        subject_ids = [
            int(item['id'])
            for item in snapshot.get('subjects', [])
            if isinstance(item, dict) and item.get('id')
        ]
        event.modifiers.set(Modifier.objects.filter(project=session.project, pk__in=modifier_ids))
        subjects = list(Subject.objects.filter(project=session.project, pk__in=subject_ids))
        event.subjects.set(subjects)
        if not event.subject_id and subjects:
            event.subject = subjects[0]
            event.save(update_fields=['subject'])
        return 'update'
    raise ValueError(_('No undo or redo information is available for this operation.'))


def serialize_event(event: ObservationEvent) -> dict:
    modifiers = list(
        event.modifiers.order_by('sort_order', 'name').values('id', 'name', 'key_binding')
    )
    subjects = [
        {
            'id': subject.id,
            'name': subject.name,
            'key_binding': subject.key_binding,
            'color': subject.color,
            'groups': [group.name for group in subject.groups.order_by('sort_order', 'name')],
        }
        for subject in event.all_subjects_ordered
    ]
    return {
        'id': event.pk,
        'behavior': event.behavior.name,
        'behavior_id': event.behavior_id,
        'behavior_mode': event.behavior.mode,
        'category': event.behavior.category.name if event.behavior.category else '',
        'color': event.behavior.color,
        'event_kind': event.event_kind,
        'timestamp_seconds': float(event.timestamp_seconds),
        'frame_index': event.frame_index,
        'comment': event.comment,
        'subject': event.subject.name if event.subject_id else '',
        'subject_id': event.subject_id,
        'subjects': subjects,
        'subjects_display': ', '.join(item['name'] for item in subjects),
        'modifiers': modifiers,
        'created_at': event.created_at.isoformat(),
    }


def serialize_annotation(annotation: SessionAnnotation) -> dict:
    return {
        'id': annotation.pk,
        'timestamp_seconds': float(annotation.timestamp_seconds),
        'title': annotation.title,
        'note': annotation.note,
        'color': annotation.color,
        'created_by': annotation.created_by.username if annotation.created_by else '',
        'created_at': annotation.created_at.isoformat(),
    }


def serialize_segment(segment: ObservationSegment) -> dict:
    return {
        'id': segment.pk,
        'title': segment.title,
        'start_seconds': float(segment.start_seconds),
        'end_seconds': float(segment.end_seconds),
        'duration_seconds': segment.duration_seconds,
        'status': segment.status,
        'assignee': segment.assignee.username if segment.assignee else '',
        'reviewer': segment.reviewer.username if segment.reviewer else '',
        'notes': segment.notes,
    }


def compute_state_status(session: ObservationSession) -> dict[int, bool]:
    state_map: dict[int, bool] = {}
    state_behaviors = {
        behavior.id for behavior in session.project.behaviors.filter(mode=Behavior.MODE_STATE)
    }
    for behavior_id in state_behaviors:
        state_map[behavior_id] = False

    event_qs = (
        ObservationEvent.objects.filter(session=session)
        .select_related('behavior')
        .order_by('timestamp_seconds', 'pk')
    )
    for event in event_qs:
        if event.behavior.mode != Behavior.MODE_STATE:
            continue
        if event.event_kind == ObservationEvent.KIND_START:
            state_map[event.behavior_id] = True
        elif event.event_kind == ObservationEvent.KIND_STOP:
            state_map[event.behavior_id] = False
    return state_map


def build_statistics(
    session: ObservationSession, duration_hint: str | float | Decimal | None = None
) -> dict:
    end_time = _session_duration(session, duration_hint)
    behaviors = list(
        session.project.behaviors.select_related('category').order_by('sort_order', 'name')
    )
    events = list(session.events.all())
    stats: dict[int, dict] = {}
    open_states: dict[int, Decimal | None] = {}

    for behavior in behaviors:
        stats[behavior.id] = {
            'behavior_id': behavior.id,
            'name': behavior.name,
            'category': behavior.category.name if behavior.category else '',
            'mode': behavior.mode,
            'color': behavior.color,
            'point_count': 0,
            'start_count': 0,
            'stop_count': 0,
            'segment_count': 0,
            'total_duration_seconds': Decimal('0'),
            'occupancy_percent': 0.0,
            'is_open': False,
        }
        open_states[behavior.id] = None

    for event in events:
        item = stats[event.behavior_id]
        if event.event_kind == ObservationEvent.KIND_POINT:
            item['point_count'] += 1
            continue
        if event.event_kind == ObservationEvent.KIND_START:
            item['start_count'] += 1
            open_states[event.behavior_id] = event.timestamp_seconds
            item['is_open'] = True
            continue
        if event.event_kind == ObservationEvent.KIND_STOP:
            item['stop_count'] += 1
            start_time = open_states.get(event.behavior_id)
            if start_time is not None and event.timestamp_seconds >= start_time:
                item['segment_count'] += 1
                item['total_duration_seconds'] += event.timestamp_seconds - start_time
            open_states[event.behavior_id] = None
            item['is_open'] = False

    for behavior in behaviors:
        start_time = open_states.get(behavior.id)
        if (
            behavior.mode == Behavior.MODE_STATE
            and start_time is not None
            and end_time >= start_time
        ):
            stats[behavior.id]['segment_count'] += 1
            stats[behavior.id]['total_duration_seconds'] += end_time - start_time
            stats[behavior.id]['is_open'] = True

    rows = []
    total_duration = Decimal('0')
    total_points = 0
    open_count = 0
    for behavior in behaviors:
        item = stats[behavior.id]
        total_duration += item['total_duration_seconds']
        total_points += item['point_count']
        if item['is_open']:
            open_count += 1
        if end_time > 0 and behavior.mode == Behavior.MODE_STATE:
            item['occupancy_percent'] = round(
                float((item['total_duration_seconds'] / end_time) * 100), 2
            )
        rows.append({**item, 'total_duration_seconds': float(item['total_duration_seconds'])})

    return {
        'session_event_count': len(events),
        'annotation_count': session.annotations.count(),
        'point_event_count': total_points,
        'open_state_count': open_count,
        'observed_span_seconds': float(end_time),
        'state_duration_seconds': float(total_duration),
        'rows': rows,
    }


def build_timeline_buckets(
    session: ObservationSession,
    duration_hint: str | float | Decimal | None = None,
    bucket_seconds: int = 60,
) -> list[dict]:
    duration = _session_duration(session, duration_hint)
    if duration <= 0:
        return []

    bucket_map: dict[int, dict] = {}
    ordered_behaviors = list(session.project.behaviors.order_by('sort_order', 'name'))
    behavior_names = {behavior.id: behavior.name for behavior in ordered_behaviors}
    for event in session.events.all():
        bucket_index = int(float(event.timestamp_seconds) // bucket_seconds)
        bucket = bucket_map.setdefault(
            bucket_index,
            {
                'index': bucket_index,
                'start_seconds': bucket_index * bucket_seconds,
                'end_seconds': (bucket_index + 1) * bucket_seconds,
                'event_count': 0,
                'point_count': 0,
                'state_change_count': 0,
                'annotation_count': 0,
                'labels': defaultdict(int),
            },
        )
        bucket['event_count'] += 1
        if event.event_kind == ObservationEvent.KIND_POINT:
            bucket['point_count'] += 1
        else:
            bucket['state_change_count'] += 1
        bucket['labels'][behavior_names[event.behavior_id]] += 1

    for annotation in session.annotations.all():
        bucket_index = int(float(annotation.timestamp_seconds) // bucket_seconds)
        bucket = bucket_map.setdefault(
            bucket_index,
            {
                'index': bucket_index,
                'start_seconds': bucket_index * bucket_seconds,
                'end_seconds': (bucket_index + 1) * bucket_seconds,
                'event_count': 0,
                'point_count': 0,
                'state_change_count': 0,
                'annotation_count': 0,
                'labels': defaultdict(int),
            },
        )
        bucket['annotation_count'] += 1
        bucket['labels'][f'Note: {annotation.title}'] += 1

    results = []
    total_buckets = int((float(duration) // bucket_seconds) + 1)
    for index in range(total_buckets):
        bucket = bucket_map.get(index)
        if bucket is None:
            bucket = {
                'index': index,
                'start_seconds': index * bucket_seconds,
                'end_seconds': (index + 1) * bucket_seconds,
                'event_count': 0,
                'point_count': 0,
                'state_change_count': 0,
                'annotation_count': 0,
                'labels': {},
            }
        else:
            bucket['labels'] = dict(
                sorted(bucket['labels'].items(), key=lambda item: (-item[1], item[0]))[:5]
            )
        results.append(bucket)
    return results


def build_track_rows(
    session: ObservationSession, duration_hint: str | float | Decimal | None = None
) -> list[dict]:
    duration = _session_duration(session, duration_hint)
    behaviors = list(
        session.project.behaviors.select_related('category').order_by('sort_order', 'name')
    )
    tracks: dict[int, dict] = {
        behavior.id: {
            'behavior_id': behavior.id,
            'name': behavior.name,
            'category': behavior.category.name if behavior.category else '',
            'mode': behavior.mode,
            'color': behavior.color,
            'segments': [],
            'points': [],
            'total_duration_seconds': 0.0,
        }
        for behavior in behaviors
    }
    open_states: dict[int, Decimal | None] = {behavior.id: None for behavior in behaviors}
    for event in session.events.all():
        track = tracks[event.behavior_id]
        if event.event_kind == ObservationEvent.KIND_POINT:
            track['points'].append(
                {
                    'event_id': event.id,
                    'seconds': float(event.timestamp_seconds),
                    'label': event.behavior.name,
                }
            )
            continue
        if event.event_kind == ObservationEvent.KIND_START:
            open_states[event.behavior_id] = event.timestamp_seconds
            continue
        start_time = open_states.get(event.behavior_id)
        if start_time is not None and event.timestamp_seconds >= start_time:
            segment = {
                'start_seconds': float(start_time),
                'end_seconds': float(event.timestamp_seconds),
                'start_event_id': next(
                    (
                        item.id
                        for item in session.events.filter(
                            behavior_id=event.behavior_id,
                            event_kind=ObservationEvent.KIND_START,
                            timestamp_seconds=start_time,
                        ).order_by('pk')
                    ),
                    None,
                ),
                'stop_event_id': event.id,
                'open': False,
            }
            track['segments'].append(segment)
            track['total_duration_seconds'] += segment['end_seconds'] - segment['start_seconds']
        open_states[event.behavior_id] = None

    for behavior in behaviors:
        start_time = open_states.get(behavior.id)
        if (
            behavior.mode == Behavior.MODE_STATE
            and start_time is not None
            and duration >= start_time
        ):
            segment = {
                'start_seconds': float(start_time),
                'end_seconds': float(duration),
                'start_event_id': next(
                    (
                        item.id
                        for item in session.events.filter(
                            behavior_id=behavior.id,
                            event_kind=ObservationEvent.KIND_START,
                            timestamp_seconds=start_time,
                        ).order_by('pk')
                    ),
                    None,
                ),
                'stop_event_id': None,
                'open': True,
            }
            tracks[behavior.id]['segments'].append(segment)
            tracks[behavior.id]['total_duration_seconds'] += (
                segment['end_seconds'] - segment['start_seconds']
            )

    return list(tracks.values())


def build_subject_statistics(
    session: ObservationSession, duration_hint: str | float | Decimal | None = None
) -> list[dict]:
    duration = _session_duration(session, duration_hint)
    rows: dict[tuple[int, int], dict] = {}
    open_states: dict[tuple[int, int], Decimal | None] = {}
    for event in session.events.all():
        related_subjects = event.all_subjects_ordered or []
        if not related_subjects:
            continue
        for subject in related_subjects:
            key = (subject.id, event.behavior_id)
            item = rows.setdefault(
                key,
                {
                    'subject': subject.name,
                    'behavior': event.behavior.name,
                    'mode': event.behavior.mode,
                    'point_count': 0,
                    'segment_count': 0,
                    'duration_seconds': Decimal('0'),
                },
            )
            open_states.setdefault(key, None)
            if event.event_kind == ObservationEvent.KIND_POINT:
                item['point_count'] += 1
            elif event.event_kind == ObservationEvent.KIND_START:
                open_states[key] = event.timestamp_seconds
            elif event.event_kind == ObservationEvent.KIND_STOP:
                start_time = open_states.get(key)
                if start_time is not None and event.timestamp_seconds >= start_time:
                    item['segment_count'] += 1
                    item['duration_seconds'] += event.timestamp_seconds - start_time
                open_states[key] = None

    for key, start_time in open_states.items():
        if start_time is not None and duration >= start_time:
            item = rows[key]
            item['segment_count'] += 1
            item['duration_seconds'] += duration - start_time

    results = []
    for item in rows.values():
        item['duration_seconds'] = round(float(item['duration_seconds']), 3)
        results.append(item)
    results.sort(key=lambda row: (row['subject'], row['behavior']))
    return results


def build_transition_rows(session: ObservationSession) -> list[dict]:
    event_rows = [event for event in session.events.all().order_by('timestamp_seconds', 'pk')]
    counters: dict[tuple[str, str], int] = defaultdict(int)
    previous_name = None
    for event in event_rows:
        current_name = event.behavior.name
        if previous_name is not None:
            counters[(previous_name, current_name)] += 1
        previous_name = current_name
    rows = [
        {'from_behavior': source, 'to_behavior': target, 'count': count}
        for (source, target), count in sorted(
            counters.items(), key=lambda item: (-item[1], item[0][0], item[0][1])
        )
    ]
    return rows


def build_audit_rows(session: ObservationSession) -> list[dict]:
    return [
        {
            'action': item.action,
            'action_label': item.get_action_display(),
            'target_type': item.target_type,
            'target_type_label': item.get_target_type_display(),
            'target_id': item.target_id,
            'actor': item.actor.username if item.actor else '',
            'summary': item.summary,
            'payload': item.payload,
            'created_at': item.created_at.isoformat(),
        }
        for item in session.audit_logs.all()
    ]


def build_interval_rows(session: ObservationSession) -> list[dict]:
    rows = []
    for behavior in session.project.behaviors.order_by('sort_order', 'name'):
        timestamps = [
            float(item.timestamp_seconds)
            for item in session.events.filter(behavior=behavior).order_by('timestamp_seconds', 'pk')
        ]
        if len(timestamps) < 2:
            rows.append(
                {
                    'name': behavior.name,
                    'category': behavior.category.name if behavior.category else '',
                    'mode': behavior.mode,
                    'interval_count': 0,
                    'mean_interval_seconds': None,
                    'min_interval_seconds': None,
                    'max_interval_seconds': None,
                }
            )
            continue
        intervals = [
            round(timestamps[i + 1] - timestamps[i], 3) for i in range(len(timestamps) - 1)
        ]
        rows.append(
            {
                'name': behavior.name,
                'category': behavior.category.name if behavior.category else '',
                'mode': behavior.mode,
                'interval_count': len(intervals),
                'mean_interval_seconds': round(sum(intervals) / len(intervals), 3),
                'min_interval_seconds': min(intervals),
                'max_interval_seconds': max(intervals),
            }
        )
    return rows


def build_integrity_report(session: ObservationSession) -> dict:
    issues = []
    open_states: dict[int, Decimal | None] = {}
    for behavior in session.project.behaviors.filter(mode=Behavior.MODE_STATE):
        open_states[behavior.id] = None

    for event in session.events.select_related('behavior').order_by('timestamp_seconds', 'pk'):
        if event.behavior.mode != Behavior.MODE_STATE:
            continue
        start_time = open_states[event.behavior_id]
        if event.event_kind == ObservationEvent.KIND_START:
            if start_time is not None:
                issues.append(
                    {
                        'severity': 'warning',
                        'message': f'Duplicate START for {event.behavior.name} at {event.timestamp_seconds}s.',
                    }
                )
            open_states[event.behavior_id] = event.timestamp_seconds
        elif event.event_kind == ObservationEvent.KIND_STOP:
            if start_time is None:
                issues.append(
                    {
                        'severity': 'warning',
                        'message': f'STOP without START for {event.behavior.name} at {event.timestamp_seconds}s.',
                    }
                )
            else:
                open_states[event.behavior_id] = None

    for behavior in session.project.behaviors.filter(mode=Behavior.MODE_STATE):
        if open_states[behavior.id] is not None:
            issues.append(
                {
                    'severity': 'warning',
                    'message': f'Open state without STOP for {behavior.name}.',
                }
            )

    return {'issue_count': len(issues), 'issues': issues}


def build_project_statistics(project: Project) -> dict:
    sessions = list(
        project.sessions.select_related('observer', 'video').prefetch_related(
            Prefetch(
                'events',
                queryset=ObservationEvent.objects.select_related('behavior', 'behavior__category')
                .prefetch_related('modifiers')
                .order_by('timestamp_seconds', 'pk'),
            ),
            'annotations',
        )
    )
    aggregate: dict[int, dict] = {}
    session_rows = []
    total_events = 0
    total_annotations = 0
    total_span = 0.0

    for behavior in project.behaviors.select_related('category').order_by('sort_order', 'name'):
        aggregate[behavior.id] = {
            'name': behavior.name,
            'category': behavior.category.name if behavior.category else '',
            'mode': behavior.mode,
            'color': behavior.color,
            'session_count': 0,
            'point_count': 0,
            'start_count': 0,
            'stop_count': 0,
            'segment_count': 0,
            'duration_seconds': 0.0,
        }

    for session in sessions:
        stats = build_statistics(session)
        total_events += stats['session_event_count']
        total_annotations += stats['annotation_count']
        total_span += stats['observed_span_seconds']
        session_rows.append(
            {
                'session_id': session.id,
                'title': session.title,
                'observer': session.observer.username if session.observer else '',
                'video': session.primary_label,
                'synced_video_count': session.video_links.count(),
                'event_count': stats['session_event_count'],
                'annotation_count': stats['annotation_count'],
                'point_event_count': stats['point_event_count'],
                'open_state_count': stats['open_state_count'],
                'observed_span_seconds': stats['observed_span_seconds'],
                'state_duration_seconds': stats['state_duration_seconds'],
            }
        )
        for row in stats['rows']:
            item = aggregate[row['behavior_id']]
            if (
                row['point_count']
                or row['segment_count']
                or row['start_count']
                or row['stop_count']
            ):
                item['session_count'] += 1
            item['point_count'] += row['point_count']
            item['start_count'] += row['start_count']
            item['stop_count'] += row['stop_count']
            item['segment_count'] += row['segment_count']
            item['duration_seconds'] += row['total_duration_seconds']

    behavior_rows = []
    for item in aggregate.values():
        item['duration_seconds'] = round(item['duration_seconds'], 3)
        behavior_rows.append(item)

    behavior_rows.sort(key=lambda row: (row['category'], row['name']))
    session_rows.sort(key=lambda row: row['title'])
    subject_rows = []
    transition_rows = []
    for session in sessions:
        subject_rows.extend(build_subject_statistics(session))
        transition_rows.extend(build_transition_rows(session))

    return {
        'project_name': project.name,
        'session_count': len(sessions),
        'video_count': project.videos.count(),
        'behavior_count': project.behaviors.count(),
        'annotation_count': total_annotations,
        'event_count': total_events,
        'observed_span_seconds': round(total_span, 3),
        'session_rows': session_rows,
        'behavior_rows': behavior_rows,
        'subject_rows': subject_rows,
        'transition_rows': transition_rows,
    }


def build_keyboard_profile_payload(project: Project) -> dict[str, dict[str, str]]:
    """Snapshot the current project shortcut assignments into serializable mappings."""
    return {
        'behavior_bindings': {
            str(item.pk): item.key_binding.upper()
            for item in project.behaviors.order_by('sort_order', 'name')
            if item.key_binding
        },
        'modifier_bindings': {
            str(item.pk): item.key_binding.upper()
            for item in project.modifiers.order_by('sort_order', 'name')
            if item.key_binding
        },
        'subject_bindings': {
            str(item.pk): item.key_binding.upper()
            for item in project.subjects.order_by('sort_order', 'name')
            if item.key_binding
        },
    }


def _bucket_signature(session: ObservationSession, bucket_seconds: int = 1) -> list[str]:
    """Create a canonical state/point signature timeline for agreement analysis."""
    duration = int(max(1, float(_session_duration(session))))
    signatures: list[str] = []
    active_states: set[str] = set()
    events = list(session.events.select_related('behavior').order_by('timestamp_seconds', 'pk'))
    index = 0
    for bucket in range(duration + 1):
        bucket_start = Decimal(str(bucket * bucket_seconds))
        bucket_end = Decimal(str((bucket + 1) * bucket_seconds))
        point_labels: list[str] = []
        while index < len(events) and events[index].timestamp_seconds < bucket_end:
            event = events[index]
            label = event.behavior.name
            if event.behavior.mode == Behavior.MODE_STATE:
                if event.event_kind == ObservationEvent.KIND_START:
                    active_states.add(label)
                elif event.event_kind == ObservationEvent.KIND_STOP and label in active_states:
                    active_states.remove(label)
            else:
                if event.timestamp_seconds >= bucket_start:
                    point_labels.append(label)
            index += 1
        current = sorted(active_states)
        point_current = sorted(point_labels)
        signature_parts = current + [f'POINT:{label}' for label in point_current]
        signatures.append(' | '.join(signature_parts) if signature_parts else '∅')
    return signatures


def build_agreement_analysis(
    reference_session: ObservationSession,
    comparison_session: ObservationSession,
    bucket_seconds: int = 1,
) -> dict:
    """Compute a simple pairwise agreement summary and confusion counts."""
    reference = _bucket_signature(reference_session, bucket_seconds=bucket_seconds)
    comparison = _bucket_signature(comparison_session, bucket_seconds=bucket_seconds)
    bucket_count = min(len(reference), len(comparison))
    if bucket_count == 0:
        return {
            'bucket_seconds': bucket_seconds,
            'bucket_count': 0,
            'percent_agreement': 0.0,
            'cohen_kappa': None,
            'confusion_rows': [],
        }
    reference = reference[:bucket_count]
    comparison = comparison[:bucket_count]
    labels = sorted(set(reference) | set(comparison))
    matches = sum(1 for left, right in zip(reference, comparison, strict=False) if left == right)
    p0 = matches / bucket_count
    ref_counts = {label: reference.count(label) for label in labels}
    cmp_counts = {label: comparison.count(label) for label in labels}
    pe = sum(
        (ref_counts[label] / bucket_count) * (cmp_counts[label] / bucket_count) for label in labels
    )
    kappa = None
    if pe < 1:
        kappa = round((p0 - pe) / (1 - pe), 4)
    confusion: dict[tuple[str, str], int] = defaultdict(int)
    for left, right in zip(reference, comparison, strict=False):
        confusion[(left, right)] += 1
    confusion_rows = [
        {'reference_label': left, 'comparison_label': right, 'count': count}
        for (left, right), count in sorted(
            confusion.items(), key=lambda item: (-item[1], item[0][0], item[0][1])
        )
    ]
    return {
        'bucket_seconds': bucket_seconds,
        'bucket_count': bucket_count,
        'percent_agreement': round(p0 * 100, 2),
        'cohen_kappa': kappa,
        'confusion_rows': confusion_rows,
    }


def build_project_boris_payload(project: Project) -> dict:
    """Build a richer BORIS-compatible project payload for exchange."""
    sessions = [
        build_boris_like_payload(session)
        for session in project.sessions.select_related('observer', 'video').prefetch_related(
            'events__behavior',
            'events__subjects',
            'events__modifiers',
            'annotations',
            'video_links__video',
        )
    ]
    payload = build_ethogram_payload(project)
    payload.update(
        {
            'schema': 'boris-project-v3',
            'subjects': [
                {
                    'name': subject.name,
                    'description': subject.description,
                    'groups': [
                        group.name for group in subject.groups.order_by('sort_order', 'name')
                    ],
                    'key_binding': subject.key_binding,
                    'color': subject.color,
                }
                for subject in project.subjects.prefetch_related('groups').order_by(
                    'sort_order', 'name'
                )
            ],
            'subject_groups': [
                {
                    'name': group.name,
                    'description': group.description,
                    'color': group.color,
                }
                for group in project.subject_groups.order_by('sort_order', 'name')
            ],
            'variables': [
                {
                    'label': definition.label,
                    'description': definition.description,
                    'value_type': definition.value_type,
                    'set_values': definition.value_options,
                    'default_value': definition.default_value,
                }
                for definition in project.variable_definitions.order_by('sort_order', 'label')
            ],
            'observation_templates': [
                {
                    'name': template.name,
                    'description': template.description,
                    'default_session_kind': template.default_session_kind,
                    'behaviors': list(
                        template.behaviors.order_by('sort_order', 'name').values_list(
                            'name', flat=True
                        )
                    ),
                    'modifiers': list(
                        template.modifiers.order_by('sort_order', 'name').values_list(
                            'name', flat=True
                        )
                    ),
                    'subjects': list(
                        template.subjects.order_by('sort_order', 'name').values_list(
                            'name', flat=True
                        )
                    ),
                    'variable_definitions': list(
                        template.variable_definitions.order_by('sort_order', 'label').values_list(
                            'label', flat=True
                        )
                    ),
                }
                for template in project.observation_templates.prefetch_related(
                    'behaviors', 'modifiers', 'subjects', 'variable_definitions'
                ).order_by('name')
            ],
            'sessions': sessions,
            'media_files': [
                {
                    'title': video.title,
                    'relative_path': _relative_media_path(video),
                    'media_kind': _media_kind_from_name(_relative_media_path(video)),
                }
                for video in project.videos.order_by('title')
            ],
        }
    )
    return payload


def build_reproducibility_bundle(project: Project) -> dict[str, bytes]:
    """Assemble a reproducible export bundle with checksums and rich metadata."""
    analytics = build_project_statistics(project)
    boris_payload = build_project_boris_payload(project)
    ethogram_payload = build_ethogram_payload(project)
    compatibility_payload = build_project_compatibility_report(project)
    files: dict[str, bytes] = {
        'ethogram.json': json.dumps(ethogram_payload, indent=2, ensure_ascii=False).encode('utf-8'),
        'analytics.json': json.dumps(analytics, indent=2, ensure_ascii=False).encode('utf-8'),
        'boris_project.json': json.dumps(boris_payload, indent=2, ensure_ascii=False).encode(
            'utf-8'
        ),
        'compatibility_report.json': json.dumps(
            compatibility_payload, indent=2, ensure_ascii=False
        ).encode('utf-8'),
    }
    session_meta = []
    for session in project.sessions.order_by('title'):
        filename = f'sessions/{slugify(session.title) or session.pk}.json'
        accessible_session = get_accessible_session(project.owner, session.pk)
        payload = build_boris_like_payload(accessible_session)
        files[filename] = json.dumps(payload, indent=2, ensure_ascii=False).encode('utf-8')
        compatibility_name = f'sessions/{slugify(session.title) or session.pk}_compatibility.json'
        files[compatibility_name] = json.dumps(
            build_session_compatibility_report(accessible_session), indent=2, ensure_ascii=False
        ).encode('utf-8')
        session_meta.append(
            {
                'id': session.pk,
                'title': session.title,
                'filename': filename,
                'compatibility_filename': compatibility_name,
                'media_paths': [
                    _relative_media_path(video)
                    for video in accessible_session.all_videos_ordered
                    if _relative_media_path(video)
                ],
            }
        )

    manifest = {
        'schema': 'pybehaviorlog-0.9.1-bundle',
        'version': '0.9.1',
        'project': {
            'name': project.name,
            'description': project.description,
            'owner': project.owner.username,
        },
        'exported_at': timezone.now().isoformat(),
        'sessions': session_meta,
        'checksums': {name: hashlib.sha256(content).hexdigest() for name, content in files.items()},
    }
    files['manifest.json'] = json.dumps(manifest, indent=2, ensure_ascii=False).encode('utf-8')
    return files


def _format_seconds_token(value: str | float | Decimal) -> str:
    """Return a stable second token for interoperability exports."""
    decimal_value = _decimal(value, default='0').quantize(Decimal('0.001'))
    token = format(decimal_value, 'f')
    return token.rstrip('0').rstrip('.') or '0'


def _build_event_interval_rows(session: ObservationSession) -> list[dict]:
    """Return normalized interval/point rows used by interoperability exports."""
    rows: list[dict] = []
    open_states: dict[tuple[int, str], dict] = {}
    end_time = _session_duration(session)
    ordered_events = list(
        session.events.select_related('behavior', 'behavior__category')
        .prefetch_related('subjects', 'subjects__groups', 'modifiers')
        .order_by('timestamp_seconds', 'pk')
    )
    for event in ordered_events:
        subjects = [subject.name for subject in event.all_subjects_ordered] or ['All subjects']
        modifiers = [modifier.name for modifier in event.modifiers.order_by('sort_order', 'name')]
        base = {
            'event_id': event.id,
            'behavior': event.behavior.name,
            'category': event.behavior.category.name if event.behavior.category else '',
            'mode': event.behavior.mode,
            'subjects': subjects,
            'modifiers': modifiers,
            'comment': event.comment,
        }
        if (
            event.event_kind == ObservationEvent.KIND_POINT
            or event.behavior.mode == Behavior.MODE_POINT
        ):
            rows.append(
                {
                    **base,
                    'event_kind': ObservationEvent.KIND_POINT,
                    'start_seconds': float(event.timestamp_seconds),
                    'end_seconds': float(event.timestamp_seconds),
                    'duration_seconds': 0.0,
                    'open': False,
                }
            )
            continue
        subject_key = '|'.join(subjects)
        state_key = (event.behavior_id, subject_key)
        if event.event_kind == ObservationEvent.KIND_START:
            open_states[state_key] = {
                **base,
                'start_seconds': float(event.timestamp_seconds),
                'start_event_id': event.id,
            }
            continue
        active = open_states.pop(state_key, None)
        if active is None:
            rows.append(
                {
                    **base,
                    'event_kind': ObservationEvent.KIND_STOP,
                    'start_seconds': float(event.timestamp_seconds),
                    'end_seconds': float(event.timestamp_seconds),
                    'duration_seconds': 0.0,
                    'open': False,
                }
            )
            continue
        rows.append(
            {
                **base,
                'event_kind': ObservationEvent.KIND_START,
                'start_seconds': active['start_seconds'],
                'end_seconds': float(event.timestamp_seconds),
                'duration_seconds': round(
                    float(event.timestamp_seconds) - active['start_seconds'], 3
                ),
                'open': False,
                'start_event_id': active['start_event_id'],
                'stop_event_id': event.id,
            }
        )
    for active in open_states.values():
        rows.append(
            {
                **active,
                'event_kind': ObservationEvent.KIND_START,
                'end_seconds': float(end_time),
                'duration_seconds': round(float(end_time) - active['start_seconds'], 3),
                'open': True,
                'stop_event_id': None,
            }
        )
    rows.sort(
        key=lambda item: (item['start_seconds'], item['behavior'], ','.join(item['subjects']))
    )
    return rows


def build_behavioral_sequences_text(session: ObservationSession, separator: str = '|') -> str:
    """Export a BORIS-style behavioral sequence text grouped by subject."""
    subject_map: dict[str, list[str]] = defaultdict(list)
    for row in _build_event_interval_rows(session):
        for subject in row['subjects'] or ['All subjects']:
            subject_map[subject].append(row['behavior'])
    lines = [f'# observation id: {session.title}', f'# project: {session.project.name}', '']
    if not subject_map:
        lines.append('No coded events')
    else:
        for subject in sorted(subject_map):
            lines.append(f'{subject}:')
            lines.append(separator.join(subject_map[subject]))
            lines.append('')
    return '\\n'.join(lines).rstrip() + '\\n'


def build_textgrid_text(session: ObservationSession) -> str:
    """Export a simple Praat TextGrid from observation rows."""
    rows = _build_event_interval_rows(session)
    end_time = max(
        float(_session_duration(session)),
        max((row['end_seconds'] for row in rows), default=0.0),
    )
    tiers: list[tuple[str, list[dict]]] = []
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        for subject in row['subjects'] or ['All subjects']:
            grouped[subject].append(row)
    for subject, subject_rows in sorted(grouped.items()):
        intervals = []
        for row in sorted(subject_rows, key=lambda item: (item['start_seconds'], item['behavior'])):
            start = row['start_seconds']
            stop = row['end_seconds']
            if stop < start:
                stop = start
            if stop == start:
                stop = round(start + 0.001, 3)
            label = row['behavior']
            if row['modifiers']:
                label = f'{label} [{", ".join(row["modifiers"])}]'
            intervals.append({'xmin': start, 'xmax': stop, 'text': label.replace('"', "'")})
        tiers.append((subject.replace('"', "'"), intervals))
    lines = [
        'File type = "ooTextFile"',
        'Object class = "TextGrid"',
        '',
        'xmin = 0',
        f'xmax = {_format_seconds_token(end_time)}',
        'tiers? <exists>',
        f'size = {len(tiers)}',
        'item []:',
    ]
    for index, (subject, intervals) in enumerate(tiers, start=1):
        lines.extend(
            [
                f'    item [{index}]:',
                '        class = "IntervalTier"',
                f'        name = "{subject}"',
                '        xmin = 0',
                f'        xmax = {_format_seconds_token(end_time)}',
                f'        intervals: size = {len(intervals)}',
            ]
        )
        for interval_index, interval in enumerate(intervals, start=1):
            lines.extend(
                [
                    f'        intervals [{interval_index}]:',
                    f'            xmin = {_format_seconds_token(interval["xmin"])}',
                    f'            xmax = {_format_seconds_token(interval["xmax"])}',
                    f'            text = "{interval["text"]}"',
                ]
            )
    return '\\n'.join(lines) + '\\n'


def build_binary_table_rows(
    session: ObservationSession, step_seconds: float = 1.0
) -> list[list[str | int]]:
    """Export a BORIS-style binary table with one column per behavior."""
    step = max(float(step_seconds), 0.1)
    duration = float(_session_duration(session))
    behaviors = list(session.project.behaviors.order_by('sort_order', 'name'))
    intervals = _build_event_interval_rows(session)
    state_rows = [row for row in intervals if row['mode'] == Behavior.MODE_STATE]
    point_rows = [row for row in intervals if row['mode'] == Behavior.MODE_POINT]
    rows: list[list[str | int]] = []
    time_value = 0.0
    while time_value <= duration + 1e-9:
        line: list[str | int] = [_format_seconds_token(time_value)]
        for behavior in behaviors:
            active = 0
            for row in state_rows:
                if (
                    row['behavior'] == behavior.name
                    and row['start_seconds'] <= time_value < row['end_seconds'] + 1e-9
                ):
                    active = 1
                    break
            if not active:
                for row in point_rows:
                    if (
                        row['behavior'] == behavior.name
                        and time_value <= row['start_seconds'] < time_value + step
                    ):
                        active = 1
                        break
            line.append(active)
        rows.append(line)
        time_value = round(time_value + step, 6)
    return rows


def _token_lookup_map(queryset, *, include_keys: bool = True) -> dict[str, object]:
    lookup: dict[str, object] = {}
    for item in queryset:
        lookup[item.name.casefold()] = item
        if include_keys and getattr(item, 'key_binding', ''):
            lookup[str(item.key_binding).casefold()] = item
    return lookup


def parse_cowlog_results_text(session: ObservationSession, raw_text: str) -> tuple[dict, dict]:
    """Parse CowLog-style plain text results into a session import payload."""
    behavior_lookup = _token_lookup_map(session.project.behaviors.all())
    modifier_lookup = _token_lookup_map(session.project.modifiers.all())
    category_lookup = {item.name.casefold() for item in session.project.categories.all()}
    lines_processed = 0
    warnings: list[str] = []
    events: list[dict] = []
    state_marker_used = False
    for line_number, raw_line in enumerate(raw_text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith('#'):
            continue
        parts = [
            part.strip()
            for part in (line.split('	') if '	' in line else line.split())
            if part.strip()
        ]
        if len(parts) < 2:
            continue
        try:
            timestamp = float(parts[0].replace(',', '.'))
        except ValueError:
            continue
        lines_processed += 1
        tokens = parts[1:]
        behavior = behavior_lookup.get(tokens[0].casefold())
        if behavior is None:
            warnings.append(
                _('Line %(line)s: unknown behavior token “%(token)s”.')
                % {'line': line_number, 'token': tokens[0]}
            )
            continue
        event_kind = ObservationEvent.KIND_POINT
        modifier_names: list[str] = []
        subject_names: list[str] = []
        for token in tokens[1:]:
            lowered = token.casefold()
            if lowered in {'point', 'start', 'stop'}:
                event_kind = lowered
                state_marker_used = True
                continue
            modifier = modifier_lookup.get(lowered)
            if modifier is not None:
                modifier_names.append(modifier.name)
                continue
            if lowered in category_lookup:
                continue
            subject_names.append(token)
        if behavior.mode == Behavior.MODE_STATE and event_kind == ObservationEvent.KIND_POINT:
            warnings.append(
                _(
                    'Line %(line)s: state behavior %(behavior)s imported as POINT because CowLog text results do not preserve paired state markers by default.'
                )
                % {'line': line_number, 'behavior': behavior.name}
            )
        events.append(
            {
                'time': timestamp,
                'behavior': behavior.name,
                'event_kind': event_kind,
                'modifiers': modifier_names,
                'subjects': subject_names,
                'comment': '',
            }
        )
    payload = {
        'schema': 'cowlog-results-v1',
        'events': events,
        'annotations': [],
    }
    report = {
        'detected_format': 'cowlog-results-v1',
        'line_count': lines_processed,
        'event_count': len(events),
        'warnings': warnings,
        'state_marker_used': state_marker_used,
    }
    return payload, report


def _normalize_import_header(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().casefold()).strip('_')


def parse_tabular_session_rows(
    session: ObservationSession, rows: list[dict[str, object]], *, source_format: str
) -> tuple[dict, dict]:
    """Parse CSV/TSV/XLSX rows with BORIS-like columns into a session payload."""
    behavior_lookup = _token_lookup_map(session.project.behaviors.all())
    modifier_lookup = _token_lookup_map(session.project.modifiers.all())
    warnings: list[str] = []
    events: list[dict] = []
    annotations: list[dict] = []
    line_count = 0
    for index, raw_row in enumerate(rows, start=2):
        row = {_normalize_import_header(key): value for key, value in raw_row.items()}
        time_token = (
            row.get('time')
            or row.get('timestamp_seconds')
            or row.get('timestamp')
            or row.get('start')
            or row.get('start_time')
            or row.get('elapsed_time')
            or row.get('media_time')
        )
        stop_token = (
            row.get('stop') or row.get('end') or row.get('stop_time') or row.get('end_time')
        )
        behavior_token = (
            row.get('behavior')
            or row.get('code')
            or row.get('behavior_code')
            or row.get('event')
            or row.get('behavior_name')
        )
        note_token = row.get('annotation') or row.get('note') or row.get('text')
        if (
            behavior_token in {None, ''}
            and note_token not in {None, ''}
            and time_token not in {None, ''}
        ):
            try:
                note_time = float(str(time_token).replace(',', '.'))
            except ValueError:
                continue
            annotations.append(
                {
                    'time': note_time,
                    'title': str(row.get('title') or 'Imported note'),
                    'note': str(note_token),
                    'color': '#f59e0b',
                }
            )
            continue
        if time_token in {None, ''} or behavior_token in {None, ''}:
            continue
        try:
            timestamp = float(str(time_token).replace(',', '.'))
        except ValueError:
            warnings.append(
                _('Row %(row)s: invalid time value “%(value)s”.')
                % {'row': index, 'value': time_token}
            )
            continue
        stop_seconds = None
        if stop_token not in {None, ''}:
            try:
                stop_seconds = float(str(stop_token).replace(',', '.'))
            except ValueError:
                stop_seconds = None
        behavior = behavior_lookup.get(str(behavior_token).casefold())
        if behavior is None:
            warnings.append(
                _('Row %(row)s: unknown behavior token “%(token)s”.')
                % {'row': index, 'token': behavior_token}
            )
            continue
        line_count += 1
        explicit_kind = _resolve_event_kind_token(
            str(
                row.get('event_kind')
                or row.get('type')
                or row.get('kind')
                or row.get('status')
                or row.get('behavior_type')
                or ''
            )
        )
        if behavior.mode == Behavior.MODE_POINT:
            event_kind = ObservationEvent.KIND_POINT
        else:
            event_kind = explicit_kind or ObservationEvent.KIND_START
        modifier_tokens = []
        for key, value in row.items():
            if key.startswith('modifier') and value not in {None, ''}:
                modifier_tokens.extend(_coerce_name_list(value))
        modifier_tokens.extend(_coerce_name_list(row.get('modifiers') or row.get('modifier')))
        normalized_modifiers = []
        for token in modifier_tokens:
            modifier = modifier_lookup.get(str(token).casefold())
            normalized_modifiers.append(modifier.name if modifier is not None else str(token))
        subject_tokens = []
        for key, value in row.items():
            if key.startswith('subject') and value not in {None, ''}:
                subject_tokens.extend(_coerce_name_list(value))
        if row.get('subject_name') not in {None, ''}:
            subject_tokens.extend(_coerce_name_list(row.get('subject_name')))
        subjects = list(dict.fromkeys(subject_tokens))
        frame_index = row.get('frame_index') or row.get('frame') or row.get('frame_number') or None
        try:
            frame_index = int(frame_index) if frame_index not in {None, ''} else None
        except (TypeError, ValueError):
            frame_index = None
        comment = str(
            row.get('comment')
            or row.get('remarks')
            or row.get('remark')
            or row.get('details')
            or row.get('image_path')
            or row.get('frame_file')
            or ''
        )
        base_event = {
            'time': timestamp,
            'behavior': behavior.name,
            'event_kind': event_kind,
            'modifiers': list(dict.fromkeys(normalized_modifiers)),
            'subjects': subjects,
            'comment': comment,
            'frame_index': frame_index,
        }
        if (
            behavior.mode == Behavior.MODE_STATE
            and stop_seconds is not None
            and stop_seconds >= timestamp
        ):
            events.append(base_event | {'event_kind': ObservationEvent.KIND_START})
            events.append(
                base_event | {'time': stop_seconds, 'event_kind': ObservationEvent.KIND_STOP}
            )
        else:
            events.append(base_event)
    payload = {
        'schema': source_format,
        'events': events,
        'annotations': annotations,
    }
    report = {
        'detected_format': source_format,
        'line_count': line_count,
        'event_count': len(events),
        'annotation_count': len(annotations),
        'warnings': warnings,
    }
    return payload, report


def parse_tabular_session_file(
    session: ObservationSession, uploaded_file, raw_bytes: bytes
) -> tuple[dict, dict]:
    """Parse CSV/TSV/XLSX tabular imports modeled on BORIS tabular exports."""
    filename = str(getattr(uploaded_file, 'name', '') or '').lower()
    if filename.endswith('.xlsx'):
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(_('The uploaded workbook is empty.'))
        headers = [_normalize_import_header(value) for value in rows[0]]
        row_dicts = []
        for row in rows[1:]:
            row_dicts.append(
                {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            )
        return parse_tabular_session_rows(session, row_dicts, source_format='boris-tabular-xlsx-v1')

    try:
        text_payload = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ValueError(_('The uploaded tabular file is not valid UTF-8 text.')) from exc
    delimiter = (
        '	'
        if ('	' in text_payload.splitlines()[0] if text_payload.splitlines() else False)
        or filename.endswith('.tsv')
        else ','
    )
    reader = csv.DictReader(io.StringIO(text_payload), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError(_('The uploaded tabular file does not contain a header row.'))
    rows = []
    for row in reader:
        rows.append({str(key): value for key, value in row.items() if key is not None})
    source_format = 'boris-tabular-tsv-v1' if delimiter == '	' else 'boris-tabular-csv-v1'
    return parse_tabular_session_rows(session, rows, source_format=source_format)


def load_session_import_payload(uploaded_file, session: ObservationSession) -> tuple[dict, dict]:
    """Load session payloads from PyBehaviorLog/BORIS JSON, tabular imports, or CowLog text exports."""
    raw_bytes = uploaded_file.read()
    report = {'warnings': []}
    filename = str(getattr(uploaded_file, 'name', '') or '').lower()
    buffer = io.BytesIO(raw_bytes)
    if zipfile.is_zipfile(buffer):
        with zipfile.ZipFile(buffer) as archive:
            names = archive.namelist()
            candidate = next((name for name in names if name.endswith('.json')), None)
            if candidate is None:
                raise ValueError(_('The uploaded archive does not contain a session JSON file.'))
            payload = json.loads(archive.read(candidate).decode('utf-8'))
            report['detected_format'] = payload.get('schema', 'json')
            report['source_name'] = candidate
            return payload, report
    if filename.endswith(('.csv', '.tsv', '.xlsx')):
        payload, parsed_report = parse_tabular_session_file(session, uploaded_file, raw_bytes)
        report.update(parsed_report)
        return payload, report
    try:
        text_payload = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ValueError(
            _('The uploaded file is not valid UTF-8 text, spreadsheet, or JSON.')
        ) from exc
    stripped = text_payload.lstrip()
    if stripped.startswith('{') or stripped.startswith('['):
        payload = json.loads(text_payload)
        report['detected_format'] = payload.get('schema', 'json')
        return payload, report
    first_line = text_payload.splitlines()[0] if text_payload.splitlines() else ''
    first_tokens = [token for token in re.split(r'[\t ]+', first_line.strip()) if token]
    if first_tokens:
        try:
            float(first_tokens[0].replace(',', '.'))
        except ValueError:
            first_token_is_time = False
        else:
            first_token_is_time = True
    else:
        first_token_is_time = False
    if first_token_is_time and filename.endswith('.txt'):
        payload, parsed_report = parse_cowlog_results_text(session, text_payload)
        report.update(parsed_report)
        return payload, report
    if ',' in first_line or '	' in first_line:
        payload, parsed_report = parse_tabular_session_file(session, uploaded_file, raw_bytes)
        report.update(parsed_report)
        return payload, report
    payload, parsed_report = parse_cowlog_results_text(session, text_payload)
    report.update(parsed_report)
    return payload, report


def build_session_compatibility_report(session: ObservationSession) -> dict:
    """Summarize what can be exchanged with BORIS and CowLog for one session."""
    stats = build_statistics(session)
    ordered_events = list(session.events.prefetch_related('modifiers', 'subjects'))
    state_event_count = sum(
        1 for event in ordered_events if event.event_kind != ObservationEvent.KIND_POINT
    )
    modifier_event_count = sum(1 for event in ordered_events if event.modifiers.exists())
    multi_subject_event_count = sum(1 for event in ordered_events if event.subjects.count() > 1)
    report = {
        'schema': 'pybehaviorlog-0.9.1-session-compatibility-report',
        'version': '0.9.1',
        'session': session.title,
        'boris': {
            'documented_exports': [
                'json',
                'behavioral_sequences',
                'textgrid',
                'binary_table',
                'csv',
                'tsv',
                'xlsx',
                'html',
                'sql',
            ],
            'documented_imports': ['json_project', 'json_observation', 'csv', 'tsv', 'xlsx'],
            'ready': True,
            'warnings': [],
        },
        'cowlog': {
            'documented_exports': ['plain_text_results'],
            'documented_imports': ['plain_text_results'],
            'ready': state_event_count == 0,
            'warnings': [],
        },
        'session_metrics': {
            'event_count': stats['session_event_count'],
            'annotation_count': stats['annotation_count'],
            'open_state_count': stats['open_state_count'],
            'state_event_count': state_event_count,
            'modifier_event_count': modifier_event_count,
            'multi_subject_event_count': multi_subject_event_count,
        },
        'certification': {
            'roundtrip_tested_families': ['boris_observation_json', 'cowlog_plain_text_results'],
            'certified_against_built_in_corpus': True,
            'fixture_version': '0.9.1',
        },
    }
    if state_event_count:
        report['cowlog']['warnings'].append(
            _(
                'CowLog plain-text exports do not preserve paired state semantics with the same fidelity as BORIS JSON.'
            )
        )
    if stats['annotation_count']:
        report['cowlog']['warnings'].append(
            _('CowLog plain-text exports do not carry annotations.')
        )
    return report


def build_project_compatibility_report(project: Project) -> dict:
    """Summarize project-level exchange coverage for BORIS and CowLog."""
    return {
        'schema': 'pybehaviorlog-0.9.1-project-compatibility-report',
        'version': '0.9.1',
        'project': project.name,
        'counts': {
            'sessions': project.sessions.count(),
            'behaviors': project.behaviors.count(),
            'subjects': project.subjects.count(),
            'modifiers': project.modifiers.count(),
            'variables': project.variable_definitions.count(),
            'templates': project.observation_templates.count(),
        },
        'supported_boris_exports': [
            'project_json',
            'session_json',
            'behavioral_sequences',
            'textgrid',
            'binary_table',
            'csv',
            'tsv',
            'xlsx',
            'html',
            'sql',
        ],
        'supported_cowlog_exports': ['plain_text_results'],
        'supported_boris_imports': ['json_project', 'json_observation', 'csv', 'tsv', 'xlsx'],
        'notes': [
            _(
                'BORIS interoperability is strongest when using the documented JSON project/observation workflows and tabular exports.'
            ),
            _(
                'CowLog interoperability currently targets the documented plain-text coding results and keyboard/behavior conventions.'
            ),
        ],
        'certification': {
            'roundtrip_tested_families': [
                'boris_project_json',
                'boris_observation_json',
                'cowlog_plain_text_results',
            ],
            'certified_against_built_in_corpus': True,
            'fixture_version': '0.9.1',
        },
        'sample_session_reports': [
            build_session_compatibility_report(session)
            for session in project.sessions.all().order_by('-created_at')[:10]
        ],
    }


def _normalize_named_item(item, default_name: str | None = None, label_mode: bool = False) -> dict:
    """Return a normalized mapping for list/dict BORIS-like import items."""
    if isinstance(item, str):
        key = 'label' if label_mode else 'name'
        return {key: item}
    if not isinstance(item, dict):
        key = 'label' if label_mode else 'name'
        return {key: default_name or str(item)}
    normalized = dict(item)
    if label_mode:
        normalized.setdefault(
            'label', normalized.get('name') or normalized.get('code') or default_name or ''
        )
        normalized.setdefault('name', normalized['label'])
    else:
        normalized.setdefault(
            'name', normalized.get('label') or normalized.get('code') or default_name or ''
        )
        normalized.setdefault('label', normalized.get('name', ''))
    return normalized


def _coerce_named_items(value, *, label_mode: bool = False) -> list[dict]:
    """Accept either a list or a mapping keyed by names and normalize it to dict items."""
    items: list[dict] = []
    if isinstance(value, dict):
        for key, item in value.items():
            items.append(_normalize_named_item(item, default_name=str(key), label_mode=label_mode))
        return items
    if isinstance(value, list):
        for item in value:
            items.append(_normalize_named_item(item, label_mode=label_mode))
    return items


def _coerce_name_list(value) -> list[str]:
    """Convert list/dict/string inputs into a flat list of names for imports."""
    if value is None:
        return []
    if isinstance(value, dict):
        if all(isinstance(item, dict) for item in value.values()):
            results = []
            for key, item in value.items():
                normalized = _normalize_named_item(item, default_name=str(key))
                results.append(normalized.get('name') or normalized.get('label') or str(key))
            return [item for item in results if item]
        return [str(key) for key, item in value.items() if item]
    if isinstance(value, list):
        results = []
        for item in value:
            if isinstance(item, dict):
                normalized = _normalize_named_item(item)
                results.append(normalized.get('name') or normalized.get('label') or '')
            else:
                results.append(str(item))
        return [item for item in results if item]
    if isinstance(value, str):
        return [part.strip() for part in re.split(r'[|,;]', value) if part.strip()]
    return [str(value)]


def _extract_observation_entries(payload: dict) -> list[dict]:
    """Extract project session/observation payloads from multiple BORIS-like shapes."""
    candidates = payload.get('sessions')
    if candidates is None:
        candidates = payload.get('observations')
    if candidates is None:
        candidates = payload.get('observation')
    if candidates is None:
        return []
    if isinstance(candidates, dict):
        rows = []
        for key, item in candidates.items():
            normalized = dict(item) if isinstance(item, dict) else {'title': str(key)}
            normalized.setdefault('title', normalized.get('description') or str(key))
            rows.append(normalized)
        return rows
    if isinstance(candidates, list):
        return [item for item in candidates if isinstance(item, dict)]
    return []


def _resolve_behavior_name(item: dict) -> str:
    return (
        item.get('behavior')
        or item.get('code')
        or item.get('behavior_code')
        or item.get('event')
        or ''
    )


def _resolve_event_kind_token(value: str | None) -> str | None:
    token = (value or '').strip().lower().replace('_', ' ')
    mapping = {
        'point': ObservationEvent.KIND_POINT,
        'instant': ObservationEvent.KIND_POINT,
        'start': ObservationEvent.KIND_START,
        'state start': ObservationEvent.KIND_START,
        'begin': ObservationEvent.KIND_START,
        'stop': ObservationEvent.KIND_STOP,
        'state stop': ObservationEvent.KIND_STOP,
        'end': ObservationEvent.KIND_STOP,
    }
    return mapping.get(token)


def _extract_media_labels(item: dict) -> list[str]:
    labels = []
    for key in (
        'synced_videos',
        'media_files',
        'media',
        'media_paths',
        'image_paths',
        'pictures',
        'frames',
    ):
        labels.extend(_coerce_name_list(item.get(key)))
    primary = item.get('primary_video') or item.get('media_file') or item.get('media_path')
    if primary:
        labels.insert(0, str(primary))
    seen = set()
    results = []
    for label in labels:
        if label and label not in seen:
            seen.add(label)
            results.append(label)
    return results


def load_project_import_payload(uploaded_file) -> tuple[dict, dict[str, dict]]:
    """Load a JSON project payload from either a JSON file or a ZIP bundle."""
    raw_bytes = uploaded_file.read()
    buffer = io.BytesIO(raw_bytes)
    bundled_sessions: dict[str, dict] = {}
    if zipfile.is_zipfile(buffer):
        with zipfile.ZipFile(buffer) as archive:
            names = archive.namelist()
            candidate = 'boris_project.json' if 'boris_project.json' in names else None
            if candidate is None:
                candidate = next(
                    (
                        name
                        for name in names
                        if name.endswith('.json') and ('project' in name or 'bundle' in name)
                    ),
                    None,
                )
            if candidate is None:
                raise ValueError(_('The uploaded archive does not contain a project JSON file.'))
            payload = json.loads(archive.read(candidate).decode('utf-8'))
            for name in names:
                if name.startswith('sessions/') and name.endswith('.json'):
                    bundled_sessions[name] = json.loads(archive.read(name).decode('utf-8'))
            return payload, bundled_sessions
    try:
        return json.loads(raw_bytes.decode('utf-8')), bundled_sessions
    except UnicodeDecodeError as exc:
        raise ValueError(_('The uploaded file is not valid UTF-8 JSON.')) from exc


@transaction.atomic
def import_project_payload(
    project: Project,
    payload: dict,
    actor,
    import_sessions: bool = True,
    create_live_sessions: bool = True,
    bundled_sessions: dict[str, dict] | None = None,
) -> dict[str, int]:
    """Import a richer BORIS-like project payload into an existing project."""
    bundled_sessions = bundled_sessions or {}
    schema = payload.get('schema')
    if schema not in {
        'boris-project-v1',
        'boris-project-v2',
        'boris-project-v3',
        'pybehaviorlog-0.8.3-bundle',
        'pybehaviorlog-0.9-bundle',
        'pybehaviorlog-0.9.1-bundle',
    }:
        raise ValueError(_('Unsupported project payload format.'))

    ethogram_payload = payload.get('ethogram') or payload
    categories_created, modifiers_created, behaviors_created = import_ethogram_payload(
        project,
        {
            **ethogram_payload,
            'schema': ethogram_payload.get('schema', 'pybehaviorlog-0.9.1-ethogram'),
        },
        replace_existing=False,
    )

    subject_group_map = {group.name: group for group in project.subject_groups.all()}
    subject_map = {subject.name: subject for subject in project.subjects.all()}
    variable_map = {item.label: item for item in project.variable_definitions.all()}
    behavior_map = {item.name: item for item in project.behaviors.all()}
    modifier_map = {item.name: item for item in project.modifiers.all()}

    subject_group_count = 0
    subject_count = 0
    variable_count = 0
    template_count = 0
    session_count = 0
    imported_event_count = 0
    imported_annotation_count = 0

    for item in _coerce_named_items(payload.get('subject_groups') or payload.get('groups')):
        group, created = SubjectGroup.objects.update_or_create(
            project=project,
            name=item.get('name') or item.get('code'),
            defaults={
                'description': item.get('description', ''),
                'color': item.get('color', '#7c3aed'),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        subject_group_map[group.name] = group
        subject_group_count += int(created)

    for item in _coerce_named_items(payload.get('subjects')):
        subject, created = Subject.objects.update_or_create(
            project=project,
            name=item.get('name') or item.get('code'),
            defaults={
                'description': item.get('description', ''),
                'key_binding': (item.get('key_binding') or '')[:1],
                'color': item.get('color', '#9333ea'),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        subject.groups.set(
            [
                subject_group_map[name]
                for name in item.get('groups', [])
                if name in subject_group_map
            ]
        )
        subject_map[subject.name] = subject
        subject_count += int(created)

    for item in _coerce_named_items(
        payload.get('variables') or payload.get('independent_variables'), label_mode=True
    ):
        definition, created = IndependentVariableDefinition.objects.update_or_create(
            project=project,
            label=item.get('label') or item.get('name') or item.get('code'),
            defaults={
                'description': item.get('description', ''),
                'value_type': item.get('value_type', IndependentVariableDefinition.TYPE_TEXT),
                'set_values': (
                    ', '.join(item.get('set_values', []))
                    if isinstance(item.get('set_values'), list)
                    else item.get('set_values', '')
                ),
                'default_value': str(item.get('default_value', '')),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        variable_map[definition.label] = definition
        variable_count += int(created)

    for item in _coerce_named_items(
        payload.get('observation_templates') or payload.get('templates')
    ):
        template, created = ObservationTemplate.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'description': item.get('description', ''),
                'default_session_kind': item.get(
                    'default_session_kind', ObservationSession.KIND_MEDIA
                ),
            },
        )
        template.behaviors.set(
            [
                behavior_map[name]
                for name in _coerce_name_list(item.get('behaviors') or item.get('codes'))
                if name in behavior_map
            ]
        )
        template.modifiers.set(
            [
                modifier_map[name]
                for name in _coerce_name_list(item.get('modifiers'))
                if name in modifier_map
            ]
        )
        template.subjects.set(
            [
                subject_map[name]
                for name in _coerce_name_list(item.get('subjects'))
                if name in subject_map
            ]
        )
        template.variable_definitions.set(
            [
                variable_map[name]
                for name in _coerce_name_list(
                    item.get('variable_definitions') or item.get('variables')
                )
                if name in variable_map
            ]
        )
        template_count += int(created)

    if import_sessions:
        session_entries = _extract_observation_entries(payload)
        if not session_entries and bundled_sessions:
            session_entries = list(bundled_sessions.values())
        for index, session_payload in enumerate(session_entries, start=1):
            observation = (session_payload.get('observations') or [{}])[0]
            title = (
                observation.get('title')
                or observation.get('description')
                or session_payload.get('session')
                or session_payload.get('title')
                or _('Imported session %(index)s') % {'index': index}
            )
            synced_titles = _extract_media_labels(observation) or _extract_media_labels(
                session_payload
            )
            primary_label = synced_titles[0] if synced_titles else ''
            existing_video = (
                project.videos.filter(title=primary_label).first() if primary_label else None
            )
            if existing_video is None and not create_live_sessions and primary_label:
                continue
            session_kind = (
                ObservationSession.KIND_MEDIA if existing_video else ObservationSession.KIND_LIVE
            )
            notes_parts = [
                item
                for item in [
                    observation.get('note'),
                    session_payload.get('review_notes'),
                ]
                if item
            ]
            if synced_titles:
                notes_parts.append(
                    _('Imported media titles: %(titles)s')
                    % {'titles': ', '.join(dict.fromkeys(synced_titles))}
                )
            session, _created = ObservationSession.objects.update_or_create(
                project=project,
                title=title,
                defaults={
                    'observer': actor,
                    'video': existing_video,
                    'session_kind': session_kind,
                    'description': session_payload.get('description', ''),
                    'notes': '\n'.join(notes_parts).strip(),
                    'review_notes': session_payload.get('review_notes', ''),
                    'workflow_status': session_payload.get(
                        'workflow_status', ObservationSession.STATUS_DRAFT
                    ),
                },
            )
            matched_videos = list(project.videos.filter(title__in=synced_titles))
            if matched_videos:
                _sync_session_videos(session, matched_videos)
            event_count, annotation_count = import_session_payload(
                session, session_payload, clear_existing=True
            )
            _log_audit(
                session,
                actor=actor,
                action=ObservationAuditLog.ACTION_IMPORT,
                target_type=ObservationAuditLog.TARGET_IMPORT,
                target_id=session.id,
                summary=f'Imported project session {session.title}.',
                payload={
                    'source_schema': schema,
                    'event_count': event_count,
                    'annotation_count': annotation_count,
                },
            )
            session_count += 1
            imported_event_count += event_count
            imported_annotation_count += annotation_count

    return {
        'categories_created': categories_created,
        'modifiers_created': modifiers_created,
        'behaviors_created': behaviors_created,
        'subject_groups_created': subject_group_count,
        'subjects_created': subject_count,
        'variables_created': variable_count,
        'templates_created': template_count,
        'sessions_imported': session_count,
        'events_imported': imported_event_count,
        'annotations_imported': imported_annotation_count,
    }


def build_ethogram_payload(project: Project) -> dict:  # pragma: no cover
    return {
        'schema': 'pybehaviorlog-0.9.1-ethogram',
        'project': {
            'name': project.name,
            'description': project.description,
            'owner': project.owner.username,
        },
        'categories': [
            {
                'name': category.name,
                'color': category.color,
                'sort_order': category.sort_order,
            }
            for category in project.categories.order_by('sort_order', 'name')
        ],
        'modifiers': [
            {
                'name': modifier.name,
                'description': modifier.description,
                'key_binding': modifier.key_binding,
                'sort_order': modifier.sort_order,
            }
            for modifier in project.modifiers.order_by('sort_order', 'name')
        ],
        'subject_groups': [
            {
                'name': group.name,
                'description': group.description,
                'color': group.color,
                'sort_order': group.sort_order,
            }
            for group in project.subject_groups.order_by('sort_order', 'name')
        ],
        'subjects': [
            {
                'name': subject.name,
                'description': subject.description,
                'key_binding': subject.key_binding,
                'color': subject.color,
                'sort_order': subject.sort_order,
                'groups': [group.name for group in subject.groups.order_by('sort_order', 'name')],
            }
            for subject in project.subjects.order_by('sort_order', 'name')
        ],
        'variables': [
            {
                'label': definition.label,
                'description': definition.description,
                'value_type': definition.value_type,
                'set_values': definition.set_values,
                'default_value': definition.default_value,
                'sort_order': definition.sort_order,
            }
            for definition in project.variable_definitions.order_by('sort_order', 'label')
        ],
        'behaviors': [
            {
                'name': behavior.name,
                'description': behavior.description,
                'key_binding': behavior.key_binding,
                'color': behavior.color,
                'mode': behavior.mode,
                'sort_order': behavior.sort_order,
                'category': behavior.category.name if behavior.category else None,
            }
            for behavior in project.behaviors.select_related('category').order_by(
                'sort_order', 'name'
            )
        ],
    }


@transaction.atomic
def import_ethogram_payload(
    project: Project, payload: dict, replace_existing: bool = False
) -> tuple[int, int, int]:  # pragma: no cover
    if payload.get('schema') not in {
        'cowlog-django-v3-ethogram',
        'cowlog-django-v4-ethogram',
        'cowlog-django-v5-ethogram',
        'pybehaviorlog-0.8-ethogram',
        'pybehaviorlog-0.8.3-ethogram',
        'pybehaviorlog-0.9-ethogram',
        'pybehaviorlog-0.9.1-ethogram',
        'boris-project-v1',
        'boris-project-v2',
        'boris-project-v3',
        'boris-observation-v1',
        'boris-observation-v2',
        'boris-observation-v3',
    }:
        raise ValueError('Unsupported JSON schema.')

    if replace_existing and (
        project.sessions.exists()
        or ObservationEvent.objects.filter(session__project=project).exists()
    ):
        raise ValueError(
            'Full replacement is blocked because the project already contains sessions or events.'
        )

    if replace_existing:
        project.behaviors.all().delete()
        project.modifiers.all().delete()
        project.categories.all().delete()
        project.subjects.all().delete()
        project.subject_groups.all().delete()
        project.variable_definitions.all().delete()

    category_map: dict[str, BehaviorCategory] = {
        category.name: category for category in project.categories.all()
    }
    subject_group_map: dict[str, SubjectGroup] = {
        group.name: group for group in project.subject_groups.all()
    }
    modifier_count = 0
    behavior_count = 0
    category_count = 0

    for item in _coerce_named_items(payload.get('categories')):
        category, created = BehaviorCategory.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'color': item.get('color', '#0f766e'),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        category_map[category.name] = category
        if created:
            category_count += 1

    for item in _coerce_named_items(payload.get('subject_groups') or payload.get('groups')):
        group, created = SubjectGroup.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'description': item.get('description', ''),
                'color': item.get('color', '#7c3aed'),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        subject_group_map[group.name] = group

    for item in _coerce_named_items(payload.get('subjects')):
        subject, _ = Subject.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'description': item.get('description', ''),
                'key_binding': (item.get('key_binding') or '')[:1].upper(),
                'color': item.get('color', '#9333ea'),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        groups = [
            subject_group_map[name] for name in item.get('groups', []) if name in subject_group_map
        ]
        subject.groups.set(groups)

    for item in _coerce_named_items(
        payload.get('variables') or payload.get('independent_variables'), label_mode=True
    ):
        IndependentVariableDefinition.objects.update_or_create(
            project=project,
            label=item.get('label') or item.get('name') or item.get('code'),
            defaults={
                'description': item.get('description', ''),
                'value_type': item.get('value_type', IndependentVariableDefinition.TYPE_TEXT),
                'set_values': (
                    ', '.join(item.get('set_values', []))
                    if isinstance(item.get('set_values'), list)
                    else item.get('set_values', '')
                ),
                'default_value': item.get('default_value', ''),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )

    for item in _coerce_named_items(payload.get('modifiers')):
        _, created = Modifier.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'description': item.get('description', ''),
                'key_binding': (item.get('key_binding') or item.get('key') or '')[0:1].upper(),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        if created:
            modifier_count += 1

    for item in _coerce_named_items(payload.get('behaviors')):
        category_name = item.get('category')
        if isinstance(category_name, dict):
            category_name = category_name.get('name')
        category = category_map.get(category_name) if category_name else None
        _, created = Behavior.objects.update_or_create(
            project=project,
            name=item['name'],
            defaults={
                'category': category,
                'description': item.get('description', ''),
                'key_binding': (item.get('key_binding') or item.get('key') or '')[0:1].upper(),
                'color': item.get('color', '#2563eb'),
                'mode': item.get('mode', Behavior.MODE_POINT),
                'sort_order': int(item.get('sort_order', 0)),
            },
        )
        if created:
            behavior_count += 1

    return category_count, modifier_count, behavior_count


def resolve_event_kind(
    session: ObservationSession, behavior: Behavior, explicit_kind: str | None
) -> str:
    if behavior.mode == Behavior.MODE_POINT:
        return ObservationEvent.KIND_POINT
    if explicit_kind in {ObservationEvent.KIND_START, ObservationEvent.KIND_STOP}:
        return explicit_kind
    is_open = False
    for event in session.events.filter(behavior=behavior).order_by('timestamp_seconds', 'pk'):
        if event.event_kind == ObservationEvent.KIND_START:
            is_open = True
        elif event.event_kind == ObservationEvent.KIND_STOP:
            is_open = False
    return ObservationEvent.KIND_STOP if is_open else ObservationEvent.KIND_START


def _sync_session_videos(session: ObservationSession, additional_videos):
    video_ids = []
    if session.video_id:
        video_ids.append(session.video_id)
    for video in additional_videos:
        if video.pk != session.video_id and video.pk not in video_ids:
            video_ids.append(video.pk)
    if not video_ids:
        SessionVideoLink.objects.filter(session=session).delete()
        return
    SessionVideoLink.objects.filter(session=session).exclude(video_id__in=video_ids).delete()
    for index, video_id in enumerate(video_ids):
        SessionVideoLink.objects.update_or_create(
            session=session,
            video_id=video_id,
            defaults={'sort_order': index},
        )


def _event_rows(session: ObservationSession):  # pragma: no cover
    linked_titles = ', '.join(video.title for video in session.all_videos_ordered)
    for event in session.events.all():
        yield [
            session.project.name,
            session.title,
            session.primary_label,
            linked_titles,
            session.observer.username if session.observer else '',
            event.behavior.category.name if event.behavior.category else '',
            event.behavior.name,
            event.behavior.mode,
            event.event_kind,
            str(event.timestamp_seconds),
            event.subjects_display,
            event.modifiers_display,
            event.comment,
            event.created_at.isoformat(),
        ]


def _annotation_rows(session: ObservationSession):  # pragma: no cover
    for annotation in session.annotations.all():
        yield [
            session.project.name,
            session.title,
            str(annotation.timestamp_seconds),
            annotation.title,
            annotation.note,
            annotation.color,
            annotation.created_by.username if annotation.created_by else '',
            annotation.created_at.isoformat(),
        ]


def _append_autosized_sheet(
    workbook: Workbook, title: str, headers: list[str], rows: list[list]
):  # pragma: no cover
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    return sheet


def _autosize_workbook(workbook: Workbook):  # pragma: no cover
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            sheet.column_dimensions[column_letter].width = min(max_len + 2, 48)


def build_boris_like_payload(session: ObservationSession) -> dict:  # pragma: no cover
    primary_video = session.all_videos_ordered[0] if session.all_videos_ordered else None
    media_paths = [
        _relative_media_path(video)
        for video in session.all_videos_ordered
        if _relative_media_path(video)
    ]
    image_paths = [path for path in media_paths if _media_kind_from_name(path) == 'image']
    picture_directory = None
    if image_paths:
        parents = {str(Path(path).parent).replace('\\', '/') for path in image_paths}
        if len(parents) == 1:
            picture_directory = next(iter(parents))
    return {
        'schema': 'boris-observation-v3',
        'project_name': session.project.name,
        'ethogram': build_ethogram_payload(session.project),
        'workflow_status': session.workflow_status,
        'review_notes': session.review_notes,
        'variables': {item.definition.label: item.value for item in session.variable_values.all()},
        'observations': [
            {
                'title': session.title,
                'primary_video': session.primary_label,
                'primary_media_path': _relative_media_path(primary_video),
                'synced_videos': [video.title for video in session.all_videos_ordered],
                'media_paths': media_paths,
                'image_paths': image_paths,
                'picture_directory': picture_directory,
                'observer': session.observer.username if session.observer else None,
                'events': [
                    {
                        'time': float(event.timestamp_seconds),
                        'behavior': event.behavior.name,
                        'event_kind': event.event_kind,
                        'modifiers': list(
                            event.modifiers.order_by('sort_order', 'name').values_list(
                                'name', flat=True
                            )
                        ),
                        'comment': event.comment,
                        'frame_index': event.frame_index,
                        'subjects': [subject.name for subject in event.all_subjects_ordered],
                    }
                    for event in session.events.all()
                ],
                'annotations': [serialize_annotation(item) for item in session.annotations.all()],
                'segments': [
                    serialize_segment(item)
                    for item in session.segments.select_related('assignee', 'reviewer').all()
                ],
            }
        ],
    }


@transaction.atomic
def import_session_payload(
    session: ObservationSession, payload: dict, clear_existing: bool = False
) -> tuple[int, int]:
    if clear_existing:
        session.events.all().delete()
        session.annotations.all().delete()
        session.segments.all().delete()

    modifier_map = {item.name: item for item in session.project.modifiers.all()}
    behavior_map = {item.name: item for item in session.project.behaviors.all()}
    subject_map = {item.name: item for item in session.project.subjects.all()}
    variable_map = {item.label: item for item in session.project.variable_definitions.all()}

    event_items: list[dict] = []
    annotation_items: list[dict] = []
    segment_items: list[dict] = []
    variable_items = payload.get('variables', {}) or payload.get('independent_variables', {}) or {}

    if payload.get('schema') in {
        'cowlog-django-v5-session',
        'pybehaviorlog-v6-session',
        'pybehaviorlog-0.8-session',
        'pybehaviorlog-0.8.3-session',
        'pybehaviorlog-0.9-session',
        'pybehaviorlog-0.9.1-session',
        'cowlog-results-v1',
        'boris-tabular-csv-v1',
        'boris-tabular-tsv-v1',
        'boris-tabular-xlsx-v1',
    }:
        event_items = payload.get('events', [])
        annotation_items = payload.get('annotations', [])
        segment_items = payload.get('segments', [])
    elif (
        payload.get('schema')
        in {
            'boris-observation-v1',
            'boris-observation-v2',
            'boris-observation-v3',
        }
        or payload.get('observations')
        or payload.get('events')
    ):
        observations = payload.get('observations', [])
        if isinstance(observations, dict):
            observations = list(observations.values())
        if observations:
            first = observations[0]
            event_items = first.get('events', [])
            annotation_items = first.get('annotations', [])
            segment_items = first.get('segments', [])
            if isinstance(first.get('variables'), dict):
                variable_items = first.get('variables')
        else:
            event_items = payload.get('events', [])
            annotation_items = payload.get('annotations', [])
            segment_items = payload.get('segments', [])
    else:
        raise ValueError(_('Unsupported session payload format.'))

    event_count = 0
    annotation_count = 0
    for raw_item in event_items:
        item = dict(raw_item) if isinstance(raw_item, dict) else {}
        behavior_name = _resolve_behavior_name(item)
        behavior = behavior_map.get(behavior_name)
        if behavior is None:
            continue
        explicit_kind = _resolve_event_kind_token(item.get('event_kind') or item.get('type'))
        event = ObservationEvent.objects.create(
            session=session,
            behavior=behavior,
            event_kind=resolve_event_kind(session, behavior, explicit_kind),
            timestamp_seconds=_decimal(
                item.get('timestamp_seconds', item.get('time', item.get('timestamp'))),
                default='0',
            ),
            frame_index=item.get('frame_index') or item.get('frame') or None,
            comment=(item.get('comment') or item.get('note') or item.get('remarks') or '').strip(),
        )
        subject_names = _coerce_name_list(item.get('subjects'))
        if item.get('subject') and item.get('subject') not in subject_names:
            subject_names = [item.get('subject'), *subject_names]
        subjects = [subject_map[name] for name in subject_names if name in subject_map]
        if subjects:
            event.subject = subjects[0]
            event.save(update_fields=['subject'])
            event.subjects.set(subjects)
        modifier_names = _coerce_name_list(item.get('modifiers') or item.get('modifier'))
        modifiers = [modifier_map[name] for name in modifier_names if name in modifier_map]
        if modifiers:
            event.modifiers.set(modifiers)
        event_count += 1

    if not isinstance(variable_items, dict):
        variable_items = {
            item.get('label') or item.get('name'): item.get('value')
            for item in _coerce_named_items(variable_items, label_mode=True)
        }
    for label, value in variable_items.items():
        definition = variable_map.get(label)
        if definition is None:
            continue
        ObservationVariableValue.objects.update_or_create(
            session=session,
            definition=definition,
            defaults={'value': str(value)},
        )

    if payload.get('workflow_status') in {
        ObservationSession.STATUS_DRAFT,
        ObservationSession.STATUS_IN_REVIEW,
        ObservationSession.STATUS_VALIDATED,
        ObservationSession.STATUS_LOCKED,
    }:
        session.workflow_status = payload['workflow_status']
        session.review_notes = payload.get('review_notes', session.review_notes or '')
        session.save(update_fields=['workflow_status', 'review_notes'])

    for raw_item in annotation_items:
        item = dict(raw_item) if isinstance(raw_item, dict) else {}
        SessionAnnotation.objects.create(
            session=session,
            timestamp_seconds=_decimal(
                item.get('timestamp_seconds', item.get('time', item.get('timestamp'))),
                default='0',
            ),
            title=(item.get('title') or 'Note').strip()[:120] or 'Note',
            note=(item.get('note') or item.get('comment') or item.get('text') or '').strip(),
            color=item.get('color', '#f59e0b'),
            created_by=session.observer,
        )
        annotation_count += 1

    member_lookup = {
        membership.user.username: membership.user
        for membership in session.project.memberships.select_related('user')
    }
    if session.project.owner_id:
        member_lookup.setdefault(session.project.owner.username, session.project.owner)
    for raw_item in segment_items:
        item = dict(raw_item) if isinstance(raw_item, dict) else {}
        status = (
            item.get('status') or ObservationSegment.STATUS_TODO
        ).strip() or ObservationSegment.STATUS_TODO
        if status not in {choice[0] for choice in ObservationSegment.STATUS_CHOICES}:
            status = ObservationSegment.STATUS_TODO
        ObservationSegment.objects.create(
            session=session,
            title=(item.get('title') or _('Imported segment')).strip()[:160]
            or _('Imported segment'),
            start_seconds=_decimal(item.get('start_seconds', item.get('start', 0)), default='0'),
            end_seconds=_decimal(
                item.get('end_seconds', item.get('end', item.get('start_seconds', 0))), default='0'
            ),
            status=status,
            assignee=member_lookup.get((item.get('assignee') or '').strip()),
            reviewer=member_lookup.get((item.get('reviewer') or '').strip()),
            notes=(item.get('notes') or item.get('note') or '').strip(),
        )

    return event_count, annotation_count


@require_GET
def healthcheck(request):
    metadata = build_release_metadata()
    return JsonResponse(
        {
            'status': 'ok',
            'service': metadata['application'],
            'version': metadata['version'],
            'time': timezone.now().isoformat(),
        }
    )


@require_GET
def release_metadata_json(request):
    return JsonResponse(build_release_metadata())


@login_required
def project_import_create(request):  # pragma: no cover
    form = ProjectImportCreateForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        uploaded = form.cleaned_data['file']
        import_sessions = form.cleaned_data['import_sessions']
        create_live_sessions = form.cleaned_data['create_live_sessions']
        try:
            payload, bundled_sessions = load_project_import_payload(uploaded)
            inferred_name = (
                (payload.get('project') or {}).get('name')
                or payload.get('name')
                or _('Imported project')
            )
            inferred_description = (
                (payload.get('project') or {}).get('description')
                or payload.get('description')
                or ''
            )
            project_name = (form.cleaned_data['name'] or inferred_name).strip()[:200] or _(
                'Imported project'
            )
            project_description = (form.cleaned_data['description'] or inferred_description).strip()
            if Project.objects.filter(owner=request.user, name=project_name).exists():
                form.add_error('name', _('You already have a project with this name.'))
            else:
                with transaction.atomic():
                    project = Project.objects.create(
                        owner=request.user,
                        name=project_name,
                        description=project_description,
                    )
                    ProjectMembership.objects.update_or_create(
                        project=project,
                        user=request.user,
                        defaults={'role': ProjectMembership.ROLE_OWNER},
                    )
                    import_project_payload(
                        project,
                        payload,
                        actor=request.user,
                        import_sessions=import_sessions,
                        create_live_sessions=create_live_sessions,
                        bundled_sessions=bundled_sessions,
                    )
                messages.success(request, _('Project created from import package.'))
                return redirect(project)
        except (json.JSONDecodeError, ValueError) as exc:
            messages.error(request, str(exc))
    return render(request, 'tracker/project_import_create.html', {'form': form})


@login_required
def project_clone(request, pk: int):  # pragma: no cover
    source = get_owned_project(request.user, pk)
    initial = {
        'name': _('%(name)s copy') % {'name': source.name},
        'description': source.description,
        'include_sessions': True,
        'include_videos': True,
    }
    form = ProjectCloneForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        if Project.objects.filter(owner=request.user, name=form.cleaned_data['name']).exists():
            form.add_error('name', _('You already have a project with this name.'))
        else:
            cloned = clone_project(
                source,
                owner=request.user,
                name=form.cleaned_data['name'],
                description=form.cleaned_data['description'],
                include_sessions=form.cleaned_data['include_sessions'],
                include_videos=form.cleaned_data['include_videos'],
            )
            messages.success(request, _('Project cloned successfully.'))
            return redirect(cloned)
    return render(request, 'tracker/project_clone_form.html', {'form': form, 'project': source})


@login_required
def home(request):  # pragma: no cover
    projects = list(
        accessible_projects_qs(request.user).prefetch_related(
            'categories',
            'modifiers',
            'behaviors',
            'videos',
            'sessions__video',
            'sessions__video_links',
            'memberships__user',
        )
    )
    for project in projects:
        project.current_role = project.role_for_user(request.user)
    return render(
        request,
        'tracker/home.html',
        {
            'projects': projects,
            'release': build_release_metadata(),
            'review_queue': build_review_queue(request.user),
        },
    )


@login_required
def project_create(request):  # pragma: no cover
    form = ProjectForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = form.save(commit=False)
        project.owner = request.user
        project.save()
        ProjectMembership.objects.update_or_create(
            project=project,
            user=request.user,
            defaults={'role': ProjectMembership.ROLE_OWNER},
        )
        messages.success(request, _('Project created successfully.'))
        return redirect(project)
    return render(request, 'tracker/project_form.html', {'form': form})


@login_required
def project_update(request, pk: int):  # pragma: no cover
    project = get_owned_project(request.user, pk)
    form = ProjectSettingsForm(request.POST or None, instance=project)
    membership_form = ProjectMembershipForm(project=project)
    keyboard_form = KeyboardProfileForm()
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Project settings updated.'))
        return redirect(project)
    memberships = project.memberships.select_related('user').order_by('role', 'user__username')
    return render(
        request,
        'tracker/project_settings.html',
        {
            'form': form,
            'project': project,
            'membership_form': membership_form,
            'memberships': memberships,
            'keyboard_form': keyboard_form,
            'keyboard_profiles': project.keyboard_profiles.order_by('name'),
        },
    )


@login_required
def project_detail(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    project = (
        accessible_projects_qs(request.user)
        .prefetch_related(
            'categories',
            'modifiers',
            'behaviors__category',
            'videos',
            'sessions__video',
            'sessions__video_links',
            'subjects__groups',
            'subject_groups',
            'variable_definitions',
            'observation_templates',
            'memberships__user',
            'keyboard_profiles',
        )
        .get(pk=project.pk)
    )
    analytics = build_project_statistics(project)
    return render(
        request,
        'tracker/project_detail.html',
        {
            'project': project,
            'is_owner': project.owner_id == request.user.id,
            'current_role': project_role(request.user, project),
            'can_edit_project': project.can_edit(request.user),
            'can_review_project': project.can_review(request.user),
            'can_manage_members': project.can_manage_members(request.user),
            'analytics': analytics,
            'memberships': project.memberships.select_related('user').order_by(
                'role', 'user__username'
            ),
            'keyboard_profiles': project.keyboard_profiles.order_by('name'),
            'open_segments': ObservationSegment.objects.filter(session__project=project)
            .exclude(status=ObservationSegment.STATUS_DONE)
            .select_related('session', 'assignee', 'reviewer')
            .order_by('session__title', 'start_seconds')[:12],
        },
    )


@login_required
def project_analytics(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    analytics = build_project_statistics(project)
    agreement = None
    comparison_session_id = request.GET.get('comparison_session')
    reference_session_id = request.GET.get('reference_session')
    if reference_session_id and comparison_session_id:
        reference_session = get_accessible_session(request.user, int(reference_session_id))
        comparison_session = get_accessible_session(request.user, int(comparison_session_id))
        if (
            reference_session.project_id == project.pk
            and comparison_session.project_id == project.pk
        ):
            agreement = build_agreement_analysis(reference_session, comparison_session)
    return render(
        request,
        'tracker/project_analytics.html',
        {
            'project': project,
            'analytics': analytics,
            'agreement': agreement,
            'session_options': project.sessions.order_by('title'),
            'reference_session_id': reference_session_id,
            'comparison_session_id': comparison_session_id,
        },
    )


@login_required
def project_export_xlsx(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    analytics = build_project_statistics(project)
    workbook = Workbook()
    overview = workbook.active
    overview.title = 'Overview'
    overview.append(['Project', project.name])
    overview.append(['Sessions', analytics['session_count']])
    overview.append(['Videos', analytics['video_count']])
    overview.append(['Behaviors', analytics['behavior_count']])
    overview.append(['Annotations', analytics['annotation_count']])
    overview.append(['Events', analytics['event_count']])
    overview.append(['Observed span seconds', analytics['observed_span_seconds']])

    _append_autosized_sheet(
        workbook,
        'Sessions',
        [
            'Session',
            'Observer',
            'Primary video',
            'Synced videos',
            'Event count',
            'Annotations',
            'Point count',
            'Open states',
            'Observed span seconds',
            'State duration seconds',
        ],
        [
            [
                row['title'],
                row['observer'],
                row['video'],
                row['synced_video_count'],
                row['event_count'],
                row['annotation_count'],
                row['point_event_count'],
                row['open_state_count'],
                row['observed_span_seconds'],
                row['state_duration_seconds'],
            ]
            for row in analytics['session_rows']
        ],
    )
    _append_autosized_sheet(
        workbook,
        'Behaviors',
        [
            'Category',
            'Behavior',
            'Mode',
            'Sessions used',
            'Point count',
            'Start count',
            'Stop count',
            'Segments',
            'Duration seconds',
        ],
        [
            [
                row['category'],
                row['name'],
                row['mode'],
                row['session_count'],
                row['point_count'],
                row['start_count'],
                row['stop_count'],
                row['segment_count'],
                row['duration_seconds'],
            ]
            for row in analytics['behavior_rows']
        ],
    )
    _append_autosized_sheet(
        workbook,
        'Subjects',
        ['Subject', 'Behavior', 'Mode', 'Point count', 'Segment count', 'Duration seconds'],
        [
            [
                row['subject'],
                row['behavior'],
                row['mode'],
                row['point_count'],
                row['segment_count'],
                row['duration_seconds'],
            ]
            for row in analytics['subject_rows']
        ],
    )
    _append_autosized_sheet(
        workbook,
        'Transitions',
        ['From behavior', 'To behavior', 'Count'],
        [
            [row['from_behavior'], row['to_behavior'], row['count']]
            for row in analytics['transition_rows']
        ],
    )
    reference_session_id = request.GET.get('reference_session')
    comparison_session_id = request.GET.get('comparison_session')
    if reference_session_id and comparison_session_id:
        reference_session = get_accessible_session(request.user, int(reference_session_id))
        comparison_session = get_accessible_session(request.user, int(comparison_session_id))
        if (
            reference_session.project_id == project.pk
            and comparison_session.project_id == project.pk
        ):
            agreement = build_agreement_analysis(reference_session, comparison_session)
            agreement_sheet = workbook.create_sheet('Agreement')
            agreement_sheet.append(['Reference session', reference_session.title])
            agreement_sheet.append(['Comparison session', comparison_session.title])
            agreement_sheet.append(['Bucket seconds', agreement['bucket_seconds']])
            agreement_sheet.append(['Bucket count', agreement['bucket_count']])
            agreement_sheet.append(['Percent agreement', agreement['percent_agreement']])
            agreement_sheet.append(['Cohen kappa', agreement['cohen_kappa']])
            agreement_sheet.append([])
            agreement_sheet.append(['Reference label', 'Comparison label', 'Count'])
            for row in agreement['confusion_rows']:
                agreement_sheet.append(
                    [row['reference_label'], row['comparison_label'], row['count']]
                )

    _autosize_workbook(workbook)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{slugify(project.name) or "project"}_analytics.xlsx"'
    )
    workbook.save(response)
    return response


@login_required
def project_export_bundle(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    bundle_files = build_reproducibility_bundle(project)
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        for name, content in bundle_files.items():
            archive.writestr(name, content)
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = (
        f'attachment; filename="{slugify(project.name) or "project"}_bundle.zip"'
    )
    return response


@login_required
def project_export_boris_json(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    payload = build_project_boris_payload(project)
    filename = f'{slugify(project.name) or "project"}_boris_project.json'
    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def project_export_compatibility_report(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    report = build_project_compatibility_report(project)
    response = HttpResponse(
        json.dumps(report, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="project_{project.pk}_compatibility_report.json"'
    )
    return response


@login_required
def project_export_ethogram(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    payload = build_ethogram_payload(project)
    filename = f'{slugify(project.name) or "project"}_ethogram.json'
    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def project_import_ethogram(request, pk: int):  # pragma: no cover
    project = get_owned_project(request.user, pk)
    form = EthogramImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        uploaded = form.cleaned_data['file']
        replace_existing = form.cleaned_data['replace_existing']
        try:
            payload = json.loads(uploaded.read().decode('utf-8'))
            category_count, modifier_count, behavior_count = import_ethogram_payload(
                project, payload, replace_existing=replace_existing
            )
        except (UnicodeDecodeError, json.JSONDecodeError):
            messages.error(request, _('The uploaded file is not valid JSON.'))
        except ValueError as exc:
            messages.error(request, str(exc))
        else:
            messages.success(
                request,
                _(
                    'Import complete. New categories: %(categories)s, modifiers: %(modifiers)s, behaviors: %(behaviors)s.'
                )
                % {
                    'categories': category_count,
                    'modifiers': modifier_count,
                    'behaviors': behavior_count,
                },
            )
            return redirect(project)
    return render(request, 'tracker/ethogram_import.html', {'form': form, 'project': project})


@login_required
def project_import_boris_json(request, pk: int):  # pragma: no cover
    project = get_owned_project(request.user, pk)
    form = ProjectBORISImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        uploaded = form.cleaned_data['file']
        import_sessions = form.cleaned_data['import_sessions']
        create_live_sessions = form.cleaned_data['create_live_sessions']
        try:
            payload, bundled_sessions = load_project_import_payload(uploaded)
            summary = import_project_payload(
                project,
                payload,
                actor=request.user,
                import_sessions=import_sessions,
                create_live_sessions=create_live_sessions,
                bundled_sessions=bundled_sessions,
            )
        except (json.JSONDecodeError, ValueError) as exc:
            messages.error(request, str(exc))
        else:
            messages.success(
                request,
                _(
                    'Project import complete. Categories: %(categories)s, modifiers: %(modifiers)s, behaviors: %(behaviors)s, subject groups: %(subject_groups)s, subjects: %(subjects)s, variables: %(variables)s, templates: %(templates)s, sessions: %(sessions)s.'
                )
                % {
                    'categories': summary['categories_created'],
                    'modifiers': summary['modifiers_created'],
                    'behaviors': summary['behaviors_created'],
                    'subject_groups': summary['subject_groups_created'],
                    'subjects': summary['subjects_created'],
                    'variables': summary['variables_created'],
                    'templates': summary['templates_created'],
                    'sessions': summary['sessions_imported'],
                },
            )
            return redirect(project)
    return render(
        request,
        'tracker/project_boris_import.html',
        {'form': form, 'project': project},
    )


@login_required
def project_membership_create(request, pk: int):  # pragma: no cover
    project = get_owned_project(request.user, pk)
    form = ProjectMembershipForm(request.POST or None, project=project)
    if request.method == 'POST' and form.is_valid():
        membership = form.save(commit=False)
        membership.project = project
        membership.save()
        messages.success(request, _('Project membership added.'))
        return redirect('tracker:project_update', pk=project.pk)
    return render(
        request,
        'tracker/project_membership_form.html',
        {'form': form, 'project': project, 'mode': 'create'},
    )


@login_required
def project_membership_update(request, pk: int):  # pragma: no cover
    membership = get_object_or_404(
        ProjectMembership.objects.select_related('project', 'user'), pk=pk
    )
    _require_project_owner(request.user, membership.project)
    form = ProjectMembershipForm(
        request.POST or None, instance=membership, project=membership.project
    )
    form.fields['user'].disabled = True
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Project membership updated.'))
        return redirect('tracker:project_update', pk=membership.project.pk)
    return render(
        request,
        'tracker/project_membership_form.html',
        {'form': form, 'project': membership.project, 'membership': membership, 'mode': 'update'},
    )


@login_required
def project_membership_delete(request, pk: int):  # pragma: no cover
    membership = get_object_or_404(
        ProjectMembership.objects.select_related('project', 'user'), pk=pk
    )
    _require_project_owner(request.user, membership.project)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = membership.project
        membership.delete()
        messages.success(request, _('Project membership deleted.'))
        return redirect('tracker:project_update', pk=project.pk)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the membership for “{membership.user.username}”',
            'project': membership.project,
        },
    )


@login_required
def keyboard_profile_create(request, pk: int):  # pragma: no cover
    project = get_owned_project(request.user, pk)
    form = KeyboardProfileForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        profile = form.save(commit=False)
        profile.project = project
        snapshot = build_keyboard_profile_payload(project)
        profile.behavior_bindings = snapshot['behavior_bindings']
        profile.modifier_bindings = snapshot['modifier_bindings']
        profile.subject_bindings = snapshot['subject_bindings']
        profile.save()
        messages.success(request, _('Keyboard profile created from the current project bindings.'))
        return redirect('tracker:project_update', pk=project.pk)
    return render(
        request,
        'tracker/keyboard_profile_form.html',
        {'form': form, 'project': project, 'mode': 'create'},
    )


@login_required
def keyboard_profile_update(request, pk: int):  # pragma: no cover
    profile = get_object_or_404(KeyboardProfile.objects.select_related('project'), pk=pk)
    _require_project_owner(request.user, profile.project)
    form = KeyboardProfileForm(request.POST or None, instance=profile)
    if request.method == 'POST' and form.is_valid():
        profile = form.save(commit=False)
        snapshot = build_keyboard_profile_payload(profile.project)
        profile.behavior_bindings = snapshot['behavior_bindings']
        profile.modifier_bindings = snapshot['modifier_bindings']
        profile.subject_bindings = snapshot['subject_bindings']
        profile.save()
        messages.success(
            request, _('Keyboard profile refreshed from the current project bindings.')
        )
        return redirect('tracker:project_update', pk=profile.project.pk)
    return render(
        request,
        'tracker/keyboard_profile_form.html',
        {'form': form, 'project': profile.project, 'profile': profile, 'mode': 'update'},
    )


@login_required
def keyboard_profile_delete(request, pk: int):  # pragma: no cover
    profile = get_object_or_404(KeyboardProfile.objects.select_related('project'), pk=pk)
    _require_project_owner(request.user, profile.project)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = profile.project
        profile.delete()
        messages.success(request, _('Keyboard profile deleted.'))
        return redirect('tracker:project_update', pk=project.pk)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the keyboard profile “{profile.name}”',
            'project': profile.project,
        },
    )


@login_required
def category_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = BehaviorCategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        category = form.save(commit=False)
        category.project = project
        category.save()
        messages.success(request, _('Category created.'))
        return redirect(project)
    return render(
        request, 'tracker/category_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def category_update(request, pk: int):  # pragma: no cover
    category = _get_owned_category(request.user, pk)
    form = BehaviorCategoryForm(request.POST or None, instance=category)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Category updated.'))
        return redirect(category.project)
    return render(
        request,
        'tracker/category_form.html',
        {'form': form, 'project': category.project, 'mode': 'update', 'object': category},
    )


@login_required
def category_delete(request, pk: int):  # pragma: no cover
    category = _get_owned_category(request.user, pk)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = category.project
        category.delete()
        messages.success(request, _('Category deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the category “{category.name}”',
            'project': category.project,
        },
    )


@login_required
def modifier_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = ModifierForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        modifier = form.save(commit=False)
        modifier.project = project
        modifier.save()
        messages.success(request, _('Modifier created.'))
        return redirect(project)
    return render(
        request, 'tracker/modifier_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def modifier_update(request, pk: int):  # pragma: no cover
    modifier = _get_owned_modifier(request.user, pk)
    form = ModifierForm(request.POST or None, instance=modifier)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Modifier updated.'))
        return redirect(modifier.project)
    return render(
        request,
        'tracker/modifier_form.html',
        {'form': form, 'project': modifier.project, 'mode': 'update', 'object': modifier},
    )


@login_required
def modifier_delete(request, pk: int):  # pragma: no cover
    modifier = _get_owned_modifier(request.user, pk)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = modifier.project
        modifier.delete()
        messages.success(request, _('Modifier deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the modifier “{modifier.name}”',
            'project': modifier.project,
        },
    )


@login_required
def subject_group_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = SubjectGroupForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        group = form.save(commit=False)
        group.project = project
        group.save()
        messages.success(request, _('Subject group created.'))
        return redirect(project)
    return render(
        request,
        'tracker/subject_group_form.html',
        {'form': form, 'project': project, 'mode': 'create'},
    )


@login_required
def subject_group_update(request, pk: int):  # pragma: no cover
    group = get_object_or_404(SubjectGroup.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user, group.project, _('You need editor permissions to edit subject groups.')
    )
    form = SubjectGroupForm(request.POST or None, instance=group)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Subject group updated.'))
        return redirect(group.project)
    return render(
        request,
        'tracker/subject_group_form.html',
        {'form': form, 'project': group.project, 'mode': 'update', 'object': group},
    )


@login_required
def subject_group_delete(request, pk: int):  # pragma: no cover
    group = get_object_or_404(SubjectGroup.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user, group.project, _('You need editor permissions to delete subject groups.')
    )
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid() and form.cleaned_data['confirm']:
        project = group.project
        group.delete()
        messages.success(request, _('Subject group deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the subject group “{group.name}”',
            'project': group.project,
        },
    )


@login_required
def subject_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = SubjectForm(request.POST or None, project=project)
    if request.method == 'POST' and form.is_valid():
        subject = form.save(commit=False)
        subject.project = project
        subject.save()
        form.save_m2m()
        messages.success(request, _('Subject created.'))
        return redirect(project)
    return render(
        request, 'tracker/subject_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def subject_update(request, pk: int):  # pragma: no cover
    subject = get_object_or_404(Subject.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user, subject.project, _('You need editor permissions to edit subjects.')
    )
    form = SubjectForm(request.POST or None, instance=subject, project=subject.project)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Subject updated.'))
        return redirect(subject.project)
    return render(
        request,
        'tracker/subject_form.html',
        {'form': form, 'project': subject.project, 'mode': 'update', 'object': subject},
    )


@login_required
def subject_delete(request, pk: int):  # pragma: no cover
    subject = get_object_or_404(Subject.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user, subject.project, _('You need editor permissions to delete subjects.')
    )
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid() and form.cleaned_data['confirm']:
        project = subject.project
        subject.delete()
        messages.success(request, _('Subject deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {'form': form, 'object_label': f'the subject “{subject.name}”', 'project': subject.project},
    )


@login_required
def independent_variable_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = IndependentVariableDefinitionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        item = form.save(commit=False)
        item.project = project
        item.save()
        messages.success(request, _('Independent variable created.'))
        return redirect(project)
    return render(
        request,
        'tracker/independent_variable_form.html',
        {'form': form, 'project': project, 'mode': 'create'},
    )


@login_required
def independent_variable_update(request, pk: int):  # pragma: no cover
    definition = get_object_or_404(
        IndependentVariableDefinition.objects.select_related('project'), pk=pk
    )
    _require_project_editor(
        request.user,
        definition.project,
        _('You need editor permissions to edit independent variables.'),
    )
    form = IndependentVariableDefinitionForm(request.POST or None, instance=definition)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Independent variable updated.'))
        return redirect(definition.project)
    return render(
        request,
        'tracker/independent_variable_form.html',
        {'form': form, 'project': definition.project, 'mode': 'update', 'object': definition},
    )


@login_required
def independent_variable_delete(request, pk: int):  # pragma: no cover
    definition = get_object_or_404(
        IndependentVariableDefinition.objects.select_related('project'), pk=pk
    )
    _require_project_editor(
        request.user,
        definition.project,
        _('You need editor permissions to delete independent variables.'),
    )
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid() and form.cleaned_data['confirm']:
        project = definition.project
        definition.delete()
        messages.success(request, _('Independent variable deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the independent variable “{definition.label}”',
            'project': definition.project,
        },
    )


@login_required
def observation_template_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = ObservationTemplateForm(request.POST or None, project=project)
    if request.method == 'POST' and form.is_valid():
        template = form.save(commit=False)
        template.project = project
        template.save()
        form.save_m2m()
        messages.success(request, _('Observation template created.'))
        return redirect(project)
    return render(
        request,
        'tracker/observation_template_form.html',
        {'form': form, 'project': project, 'mode': 'create'},
    )


@login_required
def observation_template_update(request, pk: int):  # pragma: no cover
    template = get_object_or_404(ObservationTemplate.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user,
        template.project,
        _('You need editor permissions to edit observation templates.'),
    )
    form = ObservationTemplateForm(
        request.POST or None, instance=template, project=template.project
    )
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Observation template updated.'))
        return redirect(template.project)
    return render(
        request,
        'tracker/observation_template_form.html',
        {'form': form, 'project': template.project, 'mode': 'update', 'object': template},
    )


@login_required
def observation_template_delete(request, pk: int):  # pragma: no cover
    template = get_object_or_404(ObservationTemplate.objects.select_related('project'), pk=pk)
    _require_project_editor(
        request.user,
        template.project,
        _('You need editor permissions to delete observation templates.'),
    )
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid() and form.cleaned_data['confirm']:
        project = template.project
        template.delete()
        messages.success(request, _('Observation template deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the observation template “{template.name}”',
            'project': template.project,
        },
    )


@login_required
def behavior_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = BehaviorForm(request.POST or None, project=project)
    if request.method == 'POST' and form.is_valid():
        behavior = form.save(commit=False)
        behavior.project = project
        behavior.save()
        messages.success(request, _('Behavior created.'))
        return redirect(project)
    return render(
        request, 'tracker/behavior_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def behavior_update(request, pk: int):  # pragma: no cover
    behavior = _get_owned_behavior(request.user, pk)
    form = BehaviorForm(request.POST or None, instance=behavior, project=behavior.project)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _('Behavior updated.'))
        return redirect(behavior.project)
    return render(
        request,
        'tracker/behavior_form.html',
        {'form': form, 'project': behavior.project, 'mode': 'update', 'object': behavior},
    )


@login_required
def behavior_delete(request, pk: int):  # pragma: no cover
    behavior = _get_owned_behavior(request.user, pk)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = behavior.project
        behavior.delete()
        messages.success(request, _('Behavior deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the behavior “{behavior.name}”',
            'project': behavior.project,
        },
    )


@login_required
def video_create(request, pk: int):  # pragma: no cover
    project = get_accessible_project(request.user, pk)
    _require_project_editor(request.user, project)
    form = VideoAssetForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        video = form.save(commit=False)
        video.project = project
        video.save()
        messages.success(request, _('Video added.'))
        return redirect(project)
    return render(
        request, 'tracker/video_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def video_update(request, pk: int):  # pragma: no cover
    video = _get_owned_video(request.user, pk)
    form = VideoAssetForm(request.POST or None, request.FILES or None, instance=video)
    if request.method == 'POST' and form.is_valid():
        video = form.save(commit=False)
        video.project = video.project
        video.save()
        messages.success(request, _('Video updated.'))
        return redirect(video.project)
    return render(
        request,
        'tracker/video_form.html',
        {'form': form, 'project': video.project, 'mode': 'update', 'object': video},
    )


@login_required
def video_delete(request, pk: int):  # pragma: no cover
    video = _get_owned_video(request.user, pk)
    if video.sessions.exists() or video.session_links.exists():
        messages.error(request, _('This video is still linked to one or more sessions.'))
        return redirect(video.project)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = video.project
        video.delete()
        messages.success(request, _('Video deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {'form': form, 'object_label': f'the video “{video.title}”', 'project': video.project},
    )


@login_required
def session_create(request, pk: int):  # pragma: no cover
    project = get_object_or_404(
        accessible_projects_qs(request.user).prefetch_related('videos', 'keyboard_profiles'), pk=pk
    )
    _require_project_editor(
        request.user, project, _('You need editor permissions to create sessions.')
    )
    form = ObservationSessionForm(request.POST or None, project=project)
    if request.method == 'POST' and form.is_valid():
        session = form.save(commit=False)
        session.project = project
        session.observer = request.user
        session.save()
        if session.template_id and not session.title:
            session.title = session.template.name
            session.save(update_fields=['title'])
        form.save_variable_values(session)
        _sync_session_videos(session, form.cleaned_data['additional_videos'])
        messages.success(request, _('Session created.'))
        return redirect(session)
    return render(
        request, 'tracker/session_form.html', {'form': form, 'project': project, 'mode': 'create'}
    )


@login_required
def session_update(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    _require_project_editor(
        request.user, session.project, _('You need editor permissions to update sessions.')
    )
    form = ObservationSessionForm(request.POST or None, instance=session, project=session.project)
    if request.method == 'POST' and form.is_valid():
        session = form.save()
        form.save_variable_values(session)
        _sync_session_videos(session, form.cleaned_data['additional_videos'])
        messages.success(request, _('Session updated.'))
        return redirect(session)
    return render(
        request,
        'tracker/session_form.html',
        {'form': form, 'project': session.project, 'session': session, 'mode': 'update'},
    )


@login_required
def session_delete(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    if not session.project.can_edit(request.user) and session.observer_id != request.user.id:
        raise PermissionDenied(_('Only editors or the assigned observer can delete the session.'))
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        project = session.project
        session.delete()
        messages.success(request, _('Session deleted.'))
        return redirect(project)
    return render(
        request,
        'tracker/delete_confirm.html',
        {
            'form': form,
            'object_label': f'the session “{session.title}”',
            'project': session.project,
        },
    )


@login_required
def session_import_json(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    _require_editable_session(session, request.user)
    form = SessionImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        try:
            payload, import_report = load_session_import_payload(form.cleaned_data['file'], session)
            event_count, annotation_count = import_session_payload(
                session, payload, clear_existing=form.cleaned_data['clear_existing']
            )
        except json.JSONDecodeError:
            messages.error(request, _('The uploaded JSON file is invalid.'))
        except ValueError as exc:
            messages.error(request, str(exc))
        else:
            _log_audit(
                session,
                actor=request.user,
                action=ObservationAuditLog.ACTION_IMPORT,
                target_type=ObservationAuditLog.TARGET_IMPORT,
                target_id=session.id,
                summary=f'Imported {event_count} events and {annotation_count} annotations from {import_report.get("detected_format", "unknown")}.',
                payload={
                    'event_count': event_count,
                    'annotation_count': annotation_count,
                    'import_report': import_report,
                },
            )
            for warning in import_report.get('warnings', []):
                messages.warning(request, warning)
            messages.success(
                request,
                _(
                    'Import complete from %(format)s. Imported events: %(events)s. Imported annotations: %(annotations)s.'
                )
                % {
                    'format': import_report.get('detected_format', _('unknown format')),
                    'events': event_count,
                    'annotations': annotation_count,
                },
            )
            return redirect(session)
    return render(
        request,
        'tracker/session_import.html',
        {'form': form, 'session': session, 'project': session.project},
    )


def close_open_state_events(session: ObservationSession, actor, timestamp_seconds=None) -> int:
    """Insert STOP events at the end of the session for still-open state events."""
    if timestamp_seconds is None:
        timestamp_seconds = _session_duration(session)
    stop_at = _decimal(timestamp_seconds, default='0')
    open_states: dict[int, bool] = {
        behavior.id: False
        for behavior in session.project.behaviors.filter(mode=Behavior.MODE_STATE)
    }
    for event in session.events.select_related('behavior').order_by('timestamp_seconds', 'pk'):
        if event.behavior.mode != Behavior.MODE_STATE:
            continue
        if event.event_kind == ObservationEvent.KIND_START:
            open_states[event.behavior_id] = True
        elif event.event_kind == ObservationEvent.KIND_STOP:
            open_states[event.behavior_id] = False

    created = 0
    for behavior in session.project.behaviors.filter(mode=Behavior.MODE_STATE).order_by(
        'sort_order', 'name'
    ):
        if not open_states.get(behavior.id):
            continue
        event = ObservationEvent.objects.create(
            session=session,
            behavior=behavior,
            event_kind=ObservationEvent.KIND_STOP,
            timestamp_seconds=stop_at,
            comment=_('Automatically inserted STOP for an open state.'),
        )
        _log_audit(
            session,
            actor=actor,
            action=ObservationAuditLog.ACTION_UPDATE,
            target_type=ObservationAuditLog.TARGET_EVENT,
            target_id=event.id,
            summary=f'Inserted STOP for open state {behavior.name}.',
            payload=serialize_event(event),
        )
        created += 1
    return created


@login_required
@require_POST
def session_workflow_action(request, pk: int):
    session = get_accessible_session(request.user, pk)
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else request.POST.dict()
    except json.JSONDecodeError as exc:
        return JsonResponse({'error': _('Invalid JSON: %(error)s') % {'error': exc}}, status=400)
    action = payload.get('action')
    review_notes = (payload.get('review_notes') or session.review_notes or '').strip()
    status_map = {
        'submit': ObservationSession.STATUS_IN_REVIEW,
        'validate': ObservationSession.STATUS_VALIDATED,
        'lock': ObservationSession.STATUS_LOCKED,
        'unlock': ObservationSession.STATUS_DRAFT,
        'reopen': ObservationSession.STATUS_DRAFT,
        'save_notes': None,
        'fix_unpaired_states': None,
    }
    if action not in status_map:
        return JsonResponse({'error': _('Invalid workflow action.')}, status=400)
    if action == 'submit':
        if not session.project.can_edit(request.user):
            return JsonResponse(
                {'error': _('You need editor permissions to submit a session for review.')},
                status=403,
            )
    elif action == 'fix_unpaired_states':
        if not session.project.can_edit(request.user):
            return JsonResponse(
                {'error': _('You need editor permissions to fix unpaired states.')}, status=403
            )
    else:
        if not session.project.can_review(request.user):
            return JsonResponse(
                {'error': _('You need reviewer permissions to change workflow status.')}, status=403
            )
    if action == 'fix_unpaired_states':
        fixed_count = close_open_state_events(
            session, actor=request.user, timestamp_seconds=payload.get('timestamp_seconds')
        )
        return JsonResponse(
            {
                'ok': True,
                'fixed_count': fixed_count,
                'workflow_status': session.workflow_status,
                'review_notes': session.review_notes,
            }
        )
    if action == 'save_notes':
        session.review_notes = review_notes
        session.save(update_fields=['review_notes'])
        _log_audit(
            session,
            actor=request.user,
            action=ObservationAuditLog.ACTION_UPDATE,
            target_type=ObservationAuditLog.TARGET_SESSION,
            target_id=session.id,
            summary='Review notes updated.',
            payload={'review_notes': review_notes},
        )
        return JsonResponse(
            {
                'ok': True,
                'workflow_status': session.workflow_status,
                'review_notes': session.review_notes,
            }
        )
    session.workflow_status = status_map[action]
    session.review_notes = review_notes
    now = timezone.now()
    if session.workflow_status in {
        ObservationSession.STATUS_IN_REVIEW,
        ObservationSession.STATUS_VALIDATED,
    }:
        session.reviewed_by = request.user
        session.reviewed_at = now
    if session.workflow_status == ObservationSession.STATUS_LOCKED:
        session.locked_at = now
    elif action == 'unlock':
        session.locked_at = None
    session.save(
        update_fields=['workflow_status', 'review_notes', 'reviewed_by', 'reviewed_at', 'locked_at']
    )
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_STATUS,
        target_type=ObservationAuditLog.TARGET_SESSION,
        target_id=session.id,
        summary=f'Workflow changed to {session.workflow_status}.',
        payload={'workflow_status': session.workflow_status, 'review_notes': review_notes},
    )
    return JsonResponse(
        {
            'ok': True,
            'workflow_status': session.workflow_status,
            'review_notes': session.review_notes,
        }
    )


@login_required
@require_GET
def session_audit_json(request, pk: int):
    session = get_accessible_session(request.user, pk)
    return JsonResponse({'audit_rows': build_audit_rows(session)})


@login_required
@ensure_csrf_cookie
@require_GET
def session_player(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    synced_videos = list(session.video_links.select_related('video').order_by('sort_order', 'pk'))
    if not synced_videos:
        _sync_session_videos(session, [])
        synced_videos = list(
            session.video_links.select_related('video').order_by('sort_order', 'pk')
        )
    active_profile = session.effective_keyboard_profile
    return render(
        request,
        'tracker/session_player.html',
        {
            'session': session,
            'behaviors': session.project.behaviors.select_related('category').all(),
            'modifiers': session.project.modifiers.all(),
            'subjects': session.project.subjects.prefetch_related('groups').all(),
            'subject_groups': session.project.subject_groups.all(),
            'state_status': compute_state_status(session),
            'stats': build_statistics(session),
            'timeline_buckets': build_timeline_buckets(session),
            'track_rows': build_track_rows(session),
            'subject_rows': build_subject_statistics(session),
            'transition_rows': build_transition_rows(session),
            'audit_rows': build_audit_rows(session),
            'interval_rows': build_interval_rows(session),
            'integrity_report': build_integrity_report(session),
            'annotations': session.annotations.all(),
            'variable_values': session.variable_values.all(),
            'synced_videos': synced_videos,
            'can_code_session': session.project.can_edit(request.user),
            'can_review_session': session.project.can_review(request.user),
            'active_keyboard_profile': active_profile,
            'keyboard_profiles': session.project.keyboard_profiles.order_by('name'),
            'keyboard_profile_payload': {
                'behaviors': active_profile.behavior_bindings if active_profile else {},
                'modifiers': active_profile.modifier_bindings if active_profile else {},
                'subjects': active_profile.subject_bindings if active_profile else {},
            },
            'media_analysis': build_media_analysis(session),
            'segments': session.segments.select_related('assignee', 'reviewer').order_by(
                'start_seconds', 'end_seconds'
            ),
            'segment_form': ObservationSegmentForm(project=session.project),
        },
    )


@login_required
def review_queue(request):  # pragma: no cover
    queue = build_review_queue(request.user)
    filter_name = request.GET.get('filter', 'assigned')
    rows = list(queue.get(filter_name, queue['assigned'] if request.user.is_authenticated else []))
    project_filter = request.GET.get('project', '').strip()
    status_filter = request.GET.get('status', '').strip()
    assignee_filter = request.GET.get('assignee', '').strip()
    reviewer_filter = request.GET.get('reviewer', '').strip()
    query_filter = request.GET.get('q', '').strip().lower()

    if project_filter.isdigit():
        rows = [item for item in rows if item.session.project_id == int(project_filter)]

    if status_filter == 'open':
        rows = [item for item in rows if item.status != ObservationSegment.STATUS_DONE]
    elif status_filter in {
        ObservationSegment.STATUS_TODO,
        ObservationSegment.STATUS_IN_PROGRESS,
        ObservationSegment.STATUS_DONE,
    }:
        rows = [item for item in rows if item.status == status_filter]

    if assignee_filter == 'me':
        rows = [item for item in rows if item.assignee_id == request.user.id]
    elif assignee_filter == 'unassigned':
        rows = [item for item in rows if item.assignee_id is None]

    if reviewer_filter == 'me':
        rows = [item for item in rows if item.reviewer_id == request.user.id]
    elif reviewer_filter == 'unassigned':
        rows = [item for item in rows if item.reviewer_id is None]

    if query_filter:
        rows = [
            item
            for item in rows
            if query_filter in item.title.lower()
            or query_filter in item.session.title.lower()
            or query_filter in item.session.project.name.lower()
        ]

    projects = sorted({item.session.project for item in queue['all']}, key=lambda project: project.name.lower())
    return render(
        request,
        'tracker/review_queue.html',
        {
            'queue': queue,
            'rows': rows,
            'active_filter': filter_name,
            'status_filter': status_filter,
            'project_filter': project_filter,
            'assignee_filter': assignee_filter,
            'reviewer_filter': reviewer_filter,
            'query_filter': query_filter,
            'projects': projects,
            'release': build_release_metadata(),
        },
    )


@login_required
@require_GET
def review_queue_export_segment_analytics_csv(request):
    projects = accessible_projects_qs(request.user)
    segments = (
        ObservationSegment.objects.filter(session__project__in=projects)
        .select_related('session', 'session__project', 'assignee', 'reviewer')
        .order_by('session__project__name', 'session__title', 'start_seconds', 'pk')
    )
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="review_segment_analytics.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            'project',
            'session',
            'segment',
            'status',
            'start_seconds',
            'end_seconds',
            'duration_seconds',
            'assignee',
            'reviewer',
            'notes',
        ]
    )
    for item in segments:
        writer.writerow(
            [
                item.session.project.name,
                item.session.title,
                item.title,
                item.status,
                item.start_seconds,
                item.end_seconds,
                item.duration_seconds,
                item.assignee.username if item.assignee else '',
                item.reviewer.username if item.reviewer else '',
                item.notes,
            ]
        )
    return response


@login_required
@require_POST
def segment_batch_assign(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    _require_project_reviewer(request.user, session.project)
    segment_ids = [value for value in request.POST.getlist('segment_ids') if value.isdigit()]
    if not segment_ids:
        messages.error(request, _('Select at least one review segment.'))
        return redirect(session)

    segments = list(session.segments.filter(pk__in=segment_ids))
    if not segments:
        messages.error(request, _('No matching review segments found.'))
        return redirect(session)

    member_ids = set(session.project.memberships.values_list('user_id', flat=True)) | {session.project.owner_id}

    assignee_value = request.POST.get('assignee')
    reviewer_value = request.POST.get('reviewer')
    status_value = request.POST.get('status')

    if assignee_value and (not assignee_value.isdigit() or int(assignee_value) not in member_ids):
        messages.error(request, _('Invalid assignee.'))
        return redirect(session)
    if reviewer_value and (not reviewer_value.isdigit() or int(reviewer_value) not in member_ids):
        messages.error(request, _('Invalid reviewer.'))
        return redirect(session)
    if status_value and status_value not in {
        ObservationSegment.STATUS_TODO,
        ObservationSegment.STATUS_IN_PROGRESS,
        ObservationSegment.STATUS_DONE,
    }:
        messages.error(request, _('Invalid segment status.'))
        return redirect(session)

    assignee_id = int(assignee_value) if assignee_value else None
    reviewer_id = int(reviewer_value) if reviewer_value else None
    update_assignee = request.POST.get('set_assignee') == '1'
    update_reviewer = request.POST.get('set_reviewer') == '1'
    update_status = request.POST.get('set_status') == '1'

    updated_count = 0
    for item in segments:
        changed = False
        if update_assignee and item.assignee_id != assignee_id:
            item.assignee_id = assignee_id
            changed = True
        if update_reviewer and item.reviewer_id != reviewer_id:
            item.reviewer_id = reviewer_id
            changed = True
        if update_status and item.status != status_value:
            item.status = status_value
            changed = True
        if changed:
            item.save(update_fields=['assignee', 'reviewer', 'status', 'updated_at'])
            updated_count += 1

    if updated_count:
        _log_audit(
            session,
            actor=request.user,
            action=ObservationAuditLog.ACTION_UPDATE,
            target_type=ObservationAuditLog.TARGET_SESSION,
            target_id=session.id,
            summary=f'Batch-updated {updated_count} review segments.',
            payload={
                'segment_ids': [item.id for item in segments],
                'updated_count': updated_count,
                'set_assignee': update_assignee,
                'set_reviewer': update_reviewer,
                'set_status': update_status,
            },
        )
        messages.success(request, _('Review segments updated.'))
    else:
        messages.info(request, _('No segment values changed.'))
    return redirect(session)


@login_required
def segment_create(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    _require_project_reviewer(request.user, session.project)
    form = ObservationSegmentForm(request.POST or None, project=session.project)
    if request.method == 'POST' and form.is_valid():
        segment = form.save(commit=False)
        segment.session = session
        segment.save()
        _log_audit(
            session,
            actor=request.user,
            action=ObservationAuditLog.ACTION_CREATE,
            target_type=ObservationAuditLog.TARGET_SESSION,
            target_id=segment.id,
            summary=f'Created segment {segment.title}.',
            payload={'segment_id': segment.id, 'title': segment.title},
        )
        messages.success(request, _('Review segment created.'))
        return redirect(session)
    return render(
        request,
        'tracker/segment_form.html',
        {'form': form, 'session': session, 'project': session.project, 'mode': 'create'},
    )


@login_required
def segment_update(request, pk: int):  # pragma: no cover
    segment = get_object_or_404(
        ObservationSegment.objects.select_related('session__project'), pk=pk
    )
    session = get_accessible_session(request.user, segment.session_id)
    _require_project_reviewer(request.user, session.project)
    form = ObservationSegmentForm(request.POST or None, instance=segment, project=session.project)
    if request.method == 'POST' and form.is_valid():
        form.save()
        _log_audit(
            session,
            actor=request.user,
            action=ObservationAuditLog.ACTION_UPDATE,
            target_type=ObservationAuditLog.TARGET_SESSION,
            target_id=segment.id,
            summary=f'Updated segment {segment.title}.',
            payload={'segment_id': segment.id, 'title': segment.title},
        )
        messages.success(request, _('Review segment updated.'))
        return redirect(session)
    return render(
        request,
        'tracker/segment_form.html',
        {
            'form': form,
            'session': session,
            'project': session.project,
            'mode': 'update',
            'segment': segment,
        },
    )


@login_required
def segment_delete(request, pk: int):  # pragma: no cover
    segment = get_object_or_404(
        ObservationSegment.objects.select_related('session__project'), pk=pk
    )
    session = get_accessible_session(request.user, segment.session_id)
    _require_project_reviewer(request.user, session.project)
    form = DeleteConfirmForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        title = segment.title
        segment.delete()
        _log_audit(
            session,
            actor=request.user,
            action=ObservationAuditLog.ACTION_DELETE,
            target_type=ObservationAuditLog.TARGET_SESSION,
            target_id=None,
            summary=f'Deleted segment {title}.',
            payload={'title': title},
        )
        messages.success(request, _('Review segment deleted.'))
        return redirect(session)
    return render(
        request,
        'tracker/delete_confirm.html',
        {'form': form, 'object_name': segment.title, 'cancel_url': session.get_absolute_url()},
    )


@login_required
@require_GET
def session_events_json(request, pk: int):
    session = get_accessible_session(request.user, pk)
    duration_hint = request.GET.get('duration')
    events = [serialize_event(event) for event in session.events.all()]
    return JsonResponse(
        {
            'events': events,
            'annotations': [serialize_annotation(item) for item in session.annotations.all()],
            'segments': [
                serialize_segment(item)
                for item in session.segments.select_related('assignee', 'reviewer').all()
            ],
            'state_status': compute_state_status(session),
            'stats': build_statistics(session, duration_hint=duration_hint),
            'timeline_buckets': build_timeline_buckets(session, duration_hint=duration_hint),
            'track_rows': build_track_rows(session, duration_hint=duration_hint),
            'subject_rows': build_subject_statistics(session, duration_hint=duration_hint),
            'transition_rows': build_transition_rows(session),
            'audit_rows': build_audit_rows(session),
            'interval_rows': build_interval_rows(session),
            'integrity_report': build_integrity_report(session),
            'workflow_status': session.workflow_status,
            'review_notes': session.review_notes,
            'synced_videos': [
                {
                    'id': link.video_id,
                    'title': link.video.title,
                    'url': link.video.file.url,
                    'sort_order': link.sort_order,
                    'relative_path': _relative_media_path(link.video),
                }
                for link in session.video_links.select_related('video').order_by('sort_order', 'pk')
            ],
            'media_analysis': build_media_analysis(session),
        }
    )


@login_required
@require_POST
def event_create_api(request, pk: int):
    session = get_accessible_session(request.user, pk)
    _require_editable_session(session, request.user)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError as exc:
        return JsonResponse({'error': _('Invalid JSON: %(error)s') % {'error': exc}}, status=400)

    behavior = get_object_or_404(Behavior, pk=payload.get('behavior_id'), project=session.project)
    timestamp_seconds = _decimal(payload.get('timestamp_seconds'), default='0')
    modifier_ids = payload.get('modifier_ids') or []
    subject_ids = payload.get('subject_ids') or []
    if not isinstance(modifier_ids, list):
        return JsonResponse({'error': _('modifier_ids must be a list.')}, status=400)
    if not isinstance(subject_ids, list):
        return JsonResponse({'error': _('subject_ids must be a list.')}, status=400)
    try:
        normalized_modifier_ids = [int(value) for value in modifier_ids]
        normalized_subject_ids = [int(value) for value in subject_ids]
    except (TypeError, ValueError):
        return JsonResponse({'error': _('Invalid modifier_ids or subject_ids.')}, status=400)
    modifiers = list(
        Modifier.objects.filter(project=session.project, pk__in=normalized_modifier_ids)
    )
    subjects = list(
        Subject.objects.filter(
            project=session.project, pk__in=normalized_subject_ids
        ).prefetch_related('groups')
    )
    if {modifier.pk for modifier in modifiers} != set(normalized_modifier_ids):
        return JsonResponse({'error': _('One or more modifiers are invalid.')}, status=400)
    if {subject.pk for subject in subjects} != set(normalized_subject_ids):
        return JsonResponse({'error': _('One or more subjects are invalid.')}, status=400)

    event = ObservationEvent.objects.create(
        session=session,
        behavior=behavior,
        subject=subjects[0] if subjects else None,
        event_kind=resolve_event_kind(session, behavior, payload.get('event_kind')),
        timestamp_seconds=timestamp_seconds,
        comment=(payload.get('comment') or '').strip(),
        frame_index=payload.get('frame_index') or None,
    )
    if modifiers:
        event.modifiers.set(modifiers)
    if subjects:
        event.subjects.set(subjects)
    event_snapshot = serialize_event(event)
    _push_server_history(
        request,
        session.id,
        {'target': 'event', 'action': ObservationAuditLog.ACTION_CREATE, 'after': event_snapshot},
    )
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_CREATE,
        target_type=ObservationAuditLog.TARGET_EVENT,
        target_id=event.id,
        summary=f'Created event {event.behavior.name} at {event.timestamp_seconds}s.',
        payload={'event': event_snapshot},
    )
    return JsonResponse(
        {'event': serialize_event(event), 'state_status': compute_state_status(session)}, status=201
    )


@login_required
@require_POST
def event_update_api(request, pk: int):
    event = get_object_or_404(
        ObservationEvent.objects.select_related('session__project', 'behavior'), pk=pk
    )
    session = get_accessible_session(request.user, event.session_id)
    _require_editable_session(session, request.user)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError as exc:
        return JsonResponse({'error': _('Invalid JSON: %(error)s') % {'error': exc}}, status=400)

    behavior = get_object_or_404(
        Behavior, pk=payload.get('behavior_id', event.behavior_id), project=session.project
    )
    timestamp_seconds = _decimal(
        payload.get('timestamp_seconds', event.timestamp_seconds), default='0'
    )
    modifier_ids = payload.get('modifier_ids', list(event.modifiers.values_list('id', flat=True)))
    subject_ids = payload.get(
        'subject_ids',
        list(event.subjects.values_list('id', flat=True))
        or ([event.subject_id] if event.subject_id else []),
    )
    if not isinstance(modifier_ids, list):
        return JsonResponse({'error': _('modifier_ids must be a list.')}, status=400)
    if not isinstance(subject_ids, list):
        return JsonResponse({'error': _('subject_ids must be a list.')}, status=400)
    try:
        normalized_modifier_ids = [int(value) for value in modifier_ids]
        normalized_subject_ids = [int(value) for value in subject_ids]
    except (TypeError, ValueError):
        return JsonResponse({'error': _('Invalid modifier_ids or subject_ids.')}, status=400)
    modifiers = list(
        Modifier.objects.filter(project=session.project, pk__in=normalized_modifier_ids)
    )
    subjects = list(
        Subject.objects.filter(
            project=session.project, pk__in=normalized_subject_ids
        ).prefetch_related('groups')
    )
    if {modifier.pk for modifier in modifiers} != set(normalized_modifier_ids):
        return JsonResponse({'error': _('One or more modifiers are invalid.')}, status=400)
    if {subject.pk for subject in subjects} != set(normalized_subject_ids):
        return JsonResponse({'error': _('One or more subjects are invalid.')}, status=400)

    explicit_kind = payload.get('event_kind', event.event_kind)
    if behavior.mode == Behavior.MODE_POINT:
        explicit_kind = ObservationEvent.KIND_POINT
    elif explicit_kind not in {ObservationEvent.KIND_START, ObservationEvent.KIND_STOP}:
        return JsonResponse({'error': _('Invalid event_kind for a state behavior.')}, status=400)

    before_snapshot = serialize_event(event)
    event.behavior = behavior
    event.event_kind = explicit_kind
    event.timestamp_seconds = timestamp_seconds
    event.comment = (payload.get('comment') or '').strip()
    event.frame_index = payload.get('frame_index', event.frame_index)
    event.subject = subjects[0] if subjects else None
    event.save(
        update_fields=[
            'behavior',
            'event_kind',
            'timestamp_seconds',
            'comment',
            'frame_index',
            'subject',
        ]
    )
    event.modifiers.set(modifiers)
    event.subjects.set(subjects)
    after_snapshot = serialize_event(event)
    _push_server_history(
        request,
        session.id,
        {
            'target': 'event',
            'action': ObservationAuditLog.ACTION_UPDATE,
            'before': before_snapshot,
            'after': after_snapshot,
        },
    )
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_UPDATE,
        target_type=ObservationAuditLog.TARGET_EVENT,
        target_id=event.id,
        summary=f'Updated event {event.behavior.name} at {event.timestamp_seconds}s.',
        payload={'before': before_snapshot, 'after': after_snapshot},
    )
    return JsonResponse(
        {'event': serialize_event(event), 'state_status': compute_state_status(session)}
    )


@login_required
@require_POST
def event_delete_api(request, pk: int):
    event = get_object_or_404(ObservationEvent.objects.select_related('session__project'), pk=pk)
    if event.session.project_id not in set(
        accessible_projects_qs(request.user).values_list('id', flat=True)
    ):
        raise Http404(_('Event not found.'))
    session = get_accessible_session(request.user, event.session_id)
    _require_editable_session(session, request.user)
    payload = serialize_event(event)
    event.delete()
    _push_server_history(
        request,
        session.id,
        {'target': 'event', 'action': ObservationAuditLog.ACTION_DELETE, 'before': payload},
    )
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_DELETE,
        target_type=ObservationAuditLog.TARGET_EVENT,
        target_id=payload['id'],
        summary=f'Deleted event {payload["behavior"]} at {payload["timestamp_seconds"]}s.',
        payload={'event': payload},
    )
    return JsonResponse({'ok': True, 'state_status': compute_state_status(session)})


@login_required
@require_POST
def session_undo_api(request, pk: int):
    session = get_accessible_session(request.user, pk)
    _require_editable_session(session, request.user)
    entry = _pop_server_history(request, session.id, 'undo')
    if entry is None:
        return JsonResponse({'error': _('Nothing to undo.')}, status=400)
    try:
        applied_action = _apply_history_entry(session, entry, direction='undo')
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    _push_redo_history(request, session.id, entry)
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_UPDATE,
        target_type=ObservationAuditLog.TARGET_SESSION,
        target_id=session.id,
        summary=f'Undo {applied_action} operation.',
        payload={'history_action': applied_action, 'direction': 'undo'},
    )
    return JsonResponse(
        {
            'ok': True,
            'history_action': applied_action,
            'state_status': compute_state_status(session),
        }
    )


@login_required
@require_POST
def session_redo_api(request, pk: int):
    session = get_accessible_session(request.user, pk)
    _require_editable_session(session, request.user)
    entry = _pop_server_history(request, session.id, 'redo')
    if entry is None:
        return JsonResponse({'error': _('Nothing to redo.')}, status=400)
    try:
        applied_action = _apply_history_entry(session, entry, direction='redo')
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    _restore_history_entry(request, session.id, entry, to_stack='undo')
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_UPDATE,
        target_type=ObservationAuditLog.TARGET_SESSION,
        target_id=session.id,
        summary=f'Redo {applied_action} operation.',
        payload={'history_action': applied_action, 'direction': 'redo'},
    )
    return JsonResponse(
        {
            'ok': True,
            'history_action': applied_action,
            'state_status': compute_state_status(session),
        }
    )


@login_required
@require_POST
def annotation_create_api(request, pk: int):
    session = get_accessible_session(request.user, pk)
    _require_project_reviewer(
        request.user, session.project, _('You need reviewer permissions to create annotations.')
    )
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError as exc:
        return JsonResponse({'error': _('Invalid JSON: %(error)s') % {'error': exc}}, status=400)
    annotation = SessionAnnotation.objects.create(
        session=session,
        timestamp_seconds=_decimal(payload.get('timestamp_seconds'), default='0'),
        title=(payload.get('title') or 'Note').strip()[:120] or 'Note',
        note=(payload.get('note') or '').strip(),
        color=(payload.get('color') or '#f59e0b')[:7],
        created_by=request.user,
    )
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_CREATE,
        target_type=ObservationAuditLog.TARGET_ANNOTATION,
        target_id=annotation.id,
        summary=f'Created annotation {annotation.title} at {annotation.timestamp_seconds}s.',
        payload=serialize_annotation(annotation),
    )
    return JsonResponse({'annotation': serialize_annotation(annotation)}, status=201)


@login_required
@require_POST
def annotation_update_api(request, pk: int):
    annotation = get_object_or_404(
        SessionAnnotation.objects.select_related('session__project'), pk=pk
    )
    session = get_accessible_session(request.user, annotation.session_id)
    _require_project_reviewer(
        request.user, session.project, _('You need reviewer permissions to update annotations.')
    )
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError as exc:
        return JsonResponse({'error': _('Invalid JSON: %(error)s') % {'error': exc}}, status=400)
    annotation.timestamp_seconds = _decimal(
        payload.get('timestamp_seconds', annotation.timestamp_seconds), default='0'
    )
    annotation.title = (payload.get('title') or annotation.title).strip()[:120] or 'Note'
    annotation.note = (payload.get('note') or '').strip()
    annotation.color = (payload.get('color') or annotation.color)[:7]
    annotation.save(update_fields=['timestamp_seconds', 'title', 'note', 'color'])
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_UPDATE,
        target_type=ObservationAuditLog.TARGET_ANNOTATION,
        target_id=annotation.id,
        summary=f'Updated annotation {annotation.title} at {annotation.timestamp_seconds}s.',
        payload=serialize_annotation(annotation),
    )
    return JsonResponse({'annotation': serialize_annotation(annotation), 'session_id': session.id})


@login_required
@require_POST
def annotation_delete_api(request, pk: int):
    annotation = get_object_or_404(
        SessionAnnotation.objects.select_related('session__project'), pk=pk
    )
    if annotation.session.project_id not in set(
        accessible_projects_qs(request.user).values_list('id', flat=True)
    ):
        raise Http404(_('Annotation not found.'))
    _require_editable_session(annotation.session)
    payload = serialize_annotation(annotation)
    session = annotation.session
    annotation.delete()
    _log_audit(
        session,
        actor=request.user,
        action=ObservationAuditLog.ACTION_DELETE,
        target_type=ObservationAuditLog.TARGET_ANNOTATION,
        target_id=payload['id'],
        summary=f'Deleted annotation {payload["title"]} at {payload["timestamp_seconds"]}s.',
        payload=payload,
    )
    return JsonResponse({'ok': True})


@login_required
@require_GET
def session_media_analysis_json(request, pk: int):
    """Return lightweight media diagnostics for the player UI."""
    session = get_accessible_session(request.user, pk)
    return JsonResponse({'media_analysis': build_media_analysis(session)})


@login_required
def session_export_html(request, pk: int):  # pragma: no cover
    """Export the event table as a simple standalone HTML report."""
    session = get_accessible_session(request.user, pk)
    rows = _event_rows(session)
    html = [
        '<!doctype html>',
        '<html lang="en">',
        '<head><meta charset="utf-8"><title>PyBehaviorLog session export</title>',
        '<style>body{font-family:system-ui,sans-serif;margin:2rem}table{border-collapse:collapse;width:100%}th,td{border:1px solid #cbd5e1;padding:.45rem;text-align:left}thead{background:#f8fafc}caption{text-align:left;font-weight:700;margin-bottom:.75rem}</style>',
        '</head><body>',
        f'<h1>{session.title}</h1>',
        f'<p>Project: {session.project.name}</p>',
        f'<p>Observer: {session.observer.username if session.observer else "-"}</p>',
        '<table><caption>Events</caption><thead><tr><th>Project</th><th>Session</th><th>Primary video</th><th>Synced videos</th><th>Observer</th><th>Category</th><th>Behavior</th><th>Mode</th><th>Kind</th><th>Time</th><th>Subjects</th><th>Modifiers</th><th>Comment</th><th>Created at</th></tr></thead><tbody>',
    ]
    for row in rows:
        html.append('<tr>' + ''.join(f'<td>{str(value)}</td>' for value in row) + '</tr>')
    html.append('</tbody></table></body></html>')
    response = HttpResponse(''.join(html), content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.html"'
    return response


@login_required
def session_export_sql(request, pk: int):  # pragma: no cover
    """Export session events as SQL INSERT statements for downstream analysis."""
    session = get_accessible_session(request.user, pk)
    lines = [
        '-- PyBehaviorLog 0.9.1 SQL export',
        'BEGIN;',
        'CREATE TABLE IF NOT EXISTS pybehaviorlog_event_export (project text, session text, primary_video text, synced_videos text, observer text, category text, behavior text, behavior_mode text, event_kind text, timestamp_seconds numeric(10,3), subjects text, modifiers text, comment text, created_at text);',
    ]
    for row in _event_rows(session):
        escaped = [str(value).replace("'", "''") for value in row]
        lines.append(
            'INSERT INTO pybehaviorlog_event_export (project, session, primary_video, synced_videos, observer, category, behavior, behavior_mode, event_kind, timestamp_seconds, subjects, modifiers, comment, created_at) VALUES ('
            + f"'{escaped[0]}', '{escaped[1]}', '{escaped[2]}', '{escaped[3]}', '{escaped[4]}', '{escaped[5]}', '{escaped[6]}', '{escaped[7]}', '{escaped[8]}', {escaped[9]}, '{escaped[10]}', '{escaped[11]}', '{escaped[12]}', '{escaped[13]}');"
        )
    lines.append('COMMIT;')
    response = HttpResponse('\n'.join(lines) + '\n', content_type='application/sql; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.sql"'
    return response


@login_required
def session_export_compatibility_report(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    report = build_session_compatibility_report(session)
    response = HttpResponse(
        json.dumps(report, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="session_{session.pk}_compatibility_report.json"'
    )
    return response


@login_required
def session_export_cowlog_txt(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="session_{session.pk}_cowlog_compatible.txt"'
    )
    response.write('# PyBehaviorLog 0.9.1 CowLog-compatible export\n')
    response.write(f'# session\t{session.title}\n')
    response.write(f'# project\t{session.project.name}\n')
    response.write(f'# primary_video\t{session.primary_label}\n')
    report = build_session_compatibility_report(session)
    for warning in report['cowlog']['warnings']:
        response.write(f'# warning\t{warning}\n')
    for event in session.events.all().order_by('timestamp_seconds', 'pk'):
        row = [
            _format_seconds_token(event.timestamp_seconds),
            event.behavior.name,
        ]
        modifiers = [modifier.name for modifier in event.modifiers.order_by('sort_order', 'name')]
        subjects = [subject.name for subject in event.all_subjects_ordered]
        if modifiers:
            row.extend(modifiers)
        if event.behavior.category_id:
            row.append(event.behavior.category.name)
        if event.event_kind != ObservationEvent.KIND_POINT:
            row.append(event.event_kind)
        if subjects:
            row.extend(subjects)
        response.write('\t'.join(row) + '\n')
    return response


@login_required
def session_export_behavioral_sequences(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="session_{session.pk}_behavioral_sequences.txt"'
    )
    response.write(build_behavioral_sequences_text(session))
    return response


@login_required
def session_export_textgrid(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    response = HttpResponse(content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}.TextGrid"'
    response.write(build_textgrid_text(session))
    return response


@login_required
def session_export_binary_table_tsv(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    try:
        step_seconds = float(request.GET.get('step', '1'))
    except ValueError:
        step_seconds = 1.0
    rows = build_binary_table_rows(session, step_seconds=step_seconds)
    response = HttpResponse(content_type='text/tab-separated-values; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="session_{session.pk}_binary_table.tsv"'
    )
    writer = csv.writer(response, delimiter='	')
    writer.writerow(
        [
            'time',
            *session.project.behaviors.order_by('sort_order', 'name').values_list(
                'name', flat=True
            ),
        ]
    )
    for row in rows:
        writer.writerow(row)
    return response


@login_required
def session_export_csv(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.csv"'
    response.write('﻿')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(
        [
            'project',
            'session',
            'primary_video',
            'synced_videos',
            'observer',
            'category',
            'behavior',
            'behavior_mode',
            'event_kind',
            'timestamp_seconds',
            'subjects',
            'modifiers',
            'comment',
            'created_at',
        ]
    )
    for row in _event_rows(session):
        writer.writerow(row)
    return response


@login_required
def session_export_tsv(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    response = HttpResponse(content_type='text/tab-separated-values; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.tsv"'
    writer = csv.writer(response, delimiter='\t')
    writer.writerow(['time', 'behavior', 'event_kind', 'subjects', 'modifiers', 'comment'])
    for event in session.events.all():
        writer.writerow(
            [
                str(event.timestamp_seconds),
                event.behavior.name,
                event.event_kind,
                event.subjects_display,
                event.modifiers_display,
                event.comment,
            ]
        )
    return response


@login_required
def session_export_json(request, pk: int):
    session = get_accessible_session(request.user, pk)
    payload = {
        'schema': 'pybehaviorlog-0.9.1-session',
        'project': session.project.name,
        'session': session.title,
        'video': session.primary_label,
        'primary_media_path': _relative_media_path(
            session.all_videos_ordered[0] if session.all_videos_ordered else None
        ),
        'synced_videos': [video.title for video in session.all_videos_ordered],
        'media_paths': [
            _relative_media_path(video)
            for video in session.all_videos_ordered
            if _relative_media_path(video)
        ],
        'observer': session.observer.username if session.observer else None,
        'statistics': build_statistics(session),
        'integrity_report': build_integrity_report(session),
        'interval_rows': build_interval_rows(session),
        'timeline_buckets': build_timeline_buckets(session),
        'track_rows': build_track_rows(session),
        'subject_rows': build_subject_statistics(session),
        'transition_rows': build_transition_rows(session),
        'audit_rows': build_audit_rows(session),
        'workflow_status': session.workflow_status,
        'review_notes': session.review_notes,
        'variables': {item.definition.label: item.value for item in session.variable_values.all()},
        'events': [serialize_event(event) for event in session.events.all()],
        'annotations': [serialize_annotation(item) for item in session.annotations.all()],
        'segments': [
            serialize_segment(item)
            for item in session.segments.select_related('assignee', 'reviewer').all()
        ],
    }
    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.json"'
    return response


@login_required
def session_export_boris_json(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    payload = build_boris_like_payload(session)
    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type='application/json; charset=utf-8',
    )
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_boris_like.json"'
    return response


@login_required
def session_export_xlsx(request, pk: int):  # pragma: no cover
    session = get_accessible_session(request.user, pk)
    workbook = Workbook()
    events_sheet = workbook.active
    events_sheet.title = 'Events'
    events_sheet.append(
        [
            'Project',
            'Session',
            'Primary video',
            'Synced videos',
            'Observer',
            'Category',
            'Behavior',
            'Behavior mode',
            'Event kind',
            'Timestamp seconds',
            'Modifiers',
            'Comment',
            'Created at',
        ]
    )
    for row in _event_rows(session):
        events_sheet.append(row)

    annotations_sheet = workbook.create_sheet('Annotations')
    annotations_sheet.append(
        [
            'Project',
            'Session',
            'Timestamp seconds',
            'Title',
            'Note',
            'Color',
            'Created by',
            'Created at',
        ]
    )
    for row in _annotation_rows(session):
        annotations_sheet.append(row)

    stats = build_statistics(session)
    stats_sheet = workbook.create_sheet('Summary')
    stats_sheet.append(['Session', session.title])
    stats_sheet.append(['Observed span seconds', stats['observed_span_seconds']])
    stats_sheet.append(['Event count', stats['session_event_count']])
    stats_sheet.append(['Annotation count', stats['annotation_count']])
    stats_sheet.append(['Point count', stats['point_event_count']])
    stats_sheet.append(['Open state count', stats['open_state_count']])
    stats_sheet.append(['State duration seconds', stats['state_duration_seconds']])
    stats_sheet.append(
        ['Synced videos', ', '.join(video.title for video in session.all_videos_ordered)]
    )
    stats_sheet.append([])
    stats_sheet.append(
        [
            'Category',
            'Behavior',
            'Mode',
            'Point count',
            'Start count',
            'Stop count',
            'Segments',
            'Total duration seconds',
            'Occupancy %',
            'Open',
        ]
    )
    for row in stats['rows']:
        stats_sheet.append(
            [
                row['category'],
                row['name'],
                row['mode'],
                row['point_count'],
                row['start_count'],
                row['stop_count'],
                row['segment_count'],
                row['total_duration_seconds'],
                row['occupancy_percent'],
                'yes' if row['is_open'] else 'no',
            ]
        )

    intervals_sheet = workbook.create_sheet('Intervals')
    intervals_sheet.append(
        [
            'Category',
            'Behavior',
            'Mode',
            'Interval count',
            'Mean interval seconds',
            'Min interval seconds',
            'Max interval seconds',
        ]
    )
    for row in build_interval_rows(session):
        intervals_sheet.append(
            [
                row['category'],
                row['name'],
                row['mode'],
                row['interval_count'],
                row['mean_interval_seconds'],
                row['min_interval_seconds'],
                row['max_interval_seconds'],
            ]
        )

    integrity_sheet = workbook.create_sheet('Integrity')
    integrity = build_integrity_report(session)
    integrity_sheet.append(['Issue count', integrity['issue_count']])
    integrity_sheet.append([])
    integrity_sheet.append(['Severity', 'Message'])
    for item in integrity['issues']:
        integrity_sheet.append([item['severity'], item['message']])

    buckets_sheet = workbook.create_sheet('Timeline')
    buckets_sheet.append(
        [
            'Start seconds',
            'End seconds',
            'Events',
            'Point events',
            'State changes',
            'Annotations',
            'Top labels',
        ]
    )
    for bucket in build_timeline_buckets(session):
        buckets_sheet.append(
            [
                bucket['start_seconds'],
                bucket['end_seconds'],
                bucket['event_count'],
                bucket['point_count'],
                bucket['state_change_count'],
                bucket['annotation_count'],
                ', '.join(f'{name} ({count})' for name, count in bucket['labels'].items()),
            ]
        )

    _autosize_workbook(workbook)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="session_{session.pk}_events.xlsx"'
    workbook.save(response)
    return response
